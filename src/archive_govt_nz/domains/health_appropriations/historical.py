"""Literal historical Health/GDP observations with source precision and periods."""

from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, InvalidOperation, localcontext
from io import BytesIO
from typing import TYPE_CHECKING, Any
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pyarrow as pa
from openpyxl import load_workbook

from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook
from archive_govt_nz.domains.health_appropriations.silver import (
    LINEAGE_SCHEMA,
    SILVER_SCHEMA,
)
from archive_govt_nz.domains.health_appropriations.workbook_common import (
    encode_json,
    identity,
    source_context,
    verified_snapshot,
    write_workbook_outputs,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.worksheet.worksheet import Worksheet

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_TRANSFORM = "treasury-historical-health-gdp/v1"
_MAX_NUMBER_LENGTH = 128
_MARCH = 3
_JUNE = 6
_SCHEMA = SILVER_SCHEMA.set(
    SILVER_SCHEMA.get_field_index("amount"), pa.field("amount", pa.decimal128(38, 17))
)
for _name, _type in (
    ("domain", pa.string()),
    ("year_label", pa.string()),
    ("source_number_token", pa.string()),
    ("source_number_format", pa.string()),
    ("period_end_month", pa.int32()),
    ("accounting_basis", pa.string()),
    ("valid_time_end", pa.date32()),
    ("footnotes", pa.list_(pa.string())),
):
    _SCHEMA = _SCHEMA.append(pa.field(_name, _type))
_DISPOSITIONS = pa.schema(
    [
        (key, pa.string())
        for key in (
            "source_object_sha256",
            "source_coordinate",
            "raw_value_json",
            "disposition",
            "reason",
            "record_id",
        )
    ]
)
_CONTEXT = {
    "Cash, March Years": ("Cash", 3),
    "Cash, June Years": ("Cash", 6),
    "old-GAAP": ("old-GAAP", None),
    "IFRS, June Years": ("IFRS", 6),
    "PBE Standards, June Years": ("PBE Standards", 6),
    "March Years": (None, 3),
    "June Years": (None, 6),
}


class _NoDTD(ET.TreeBuilder):
    def doctype(self, _name: str, _pubid: str | None, _system: str | None) -> None:
        """Reject declarations before any entity expansion can occur."""
        raise ValueError("xml_doctype_forbidden")


def _xml(package: ZipFile, member: str) -> ET.Element:
    # A custom TreeBuilder rejects every DTD, including UTF-16 declarations.
    # Workbook package/cell limits are enforced before this inert lexical read.
    return ET.fromstring(package.read(member), parser=ET.XMLParser(target=_NoDTD()))  # noqa: S314


def _number_tokens(payload: bytes) -> dict[str, dict[str, str]]:
    # Caller has already applied workbook member, expansion and cell-scan caps.
    with ZipFile(BytesIO(payload)) as package:
        relations = list(_xml(package, "xl/_rels/workbook.xml.rels"))
        by_id = {item.get("Id"): item for item in relations}
        if len(by_id) != len(relations):
            raise ValueError("duplicate_sheet_relationship")
        result = {}
        for sheet in _xml(package, "xl/workbook.xml").findall(
            f"{_NS}sheets/{_NS}sheet"
        ):
            relation = by_id[sheet.get(f"{_REL}id")]
            target = relation.get("Target", "")
            member = target[1:] if target.startswith("/") else "xl/" + target
            if (
                relation.get("TargetMode") == "External"
                or re.fullmatch(r"xl/worksheets/[A-Za-z0-9_.-]+\.xml", member) is None
            ):
                raise ValueError("unsupported_sheet_relationship")
            tokens = {}
            seen = set()
            for cell in _xml(package, member).findall(
                f"{_NS}sheetData/{_NS}row/{_NS}c"
            ):
                coordinate = cell.get("r", "")
                if (
                    re.fullmatch(r"[A-Z]+[1-9][0-9]*", coordinate) is None
                    or coordinate in seen
                ):
                    raise ValueError("ambiguous_source_cell")
                seen.add(coordinate)
                values = cell.findall(f"{_NS}v")
                if len(values) > 1:
                    raise ValueError("ambiguous_source_value")
                if (
                    cell.get("t", "n") == "n"
                    and cell.find(f"{_NS}f") is None
                    and values
                    and values[0].text is not None
                ):
                    tokens[coordinate] = values[0].text
            name = sheet.get("name", "")
            if not name or name in result:
                raise ValueError("ambiguous_sheet_name")
            result[name] = tokens
        return result


def _exact_amount(token: str) -> Decimal | None:
    if len(token) > _MAX_NUMBER_LENGTH:
        return None
    try:
        with localcontext() as context:
            context.prec = 38
            number = Decimal(token)
            if not number.is_finite() or number.copy_abs() >= Decimal("1e21"):
                return None
            scaled = number.quantize(Decimal("1e-17"))
            return scaled if scaled == number else None
    except InvalidOperation:
        return None


def _coordinate(sheet: Worksheet, cell: str) -> str:
    return f"'{sheet.title}'!{cell}"


def _notes(sheet: Worksheet, end: int) -> dict[str, tuple[str, str]]:
    notes: dict[str, tuple[str, str]] = {}
    for row in range(5, end):
        text = sheet.cell(row, 1).value
        if isinstance(text, str) and re.match(r"[†*^#] ", text):
            if text[0] in notes:
                raise ValueError("ambiguous_footnote")
            notes[text[0]] = (f"A{row}", text)
    return notes


def _year_rows(sheet: Worksheet, end: int) -> list[tuple[int, int, str, str]]:
    rows = []
    previous = None
    ended = False
    for row in range(5, end):
        raw_year = sheet.cell(row, 2).value
        match = re.fullmatch(r"([1-9][0-9]{3})([†*^#]*)", str(raw_year))
        if match is None:
            if raw_year is not None and not ended:
                raise ValueError("historical_year")
            ended = True
            continue
        year, markers = int(match[1]), match[2]
        if (
            ended
            or (previous is not None and year != previous + 1)
            or len(set(markers)) != len(markers)
        ):
            raise ValueError("historical_year_sequence")
        previous = year
        rows.append((row, year, markers, str(raw_year)))
    if previous is None:
        raise ValueError("empty_historical_series")
    return rows


def _period(sheet: Worksheet, row: int, current: dict[str, Any]) -> dict[str, Any]:
    label = sheet.cell(row, 1).value
    if label is None:
        return current
    key = str(label).strip()
    health = sheet.title == "Spending"
    if key not in _CONTEXT or (health == (key in ("March Years", "June Years"))):
        raise ValueError("historical_context")
    basis, month = _CONTEXT[key]
    result = {**current, "accounting_basis": basis, "basis_cell": f"A{row}"}
    if month is not None:
        result.update(period_end_month=month, period_cell=f"A{row}")
    elif current.get("period_end_month") != _JUNE:
        raise ValueError("historical_gaap_period")
    return result


def _inputs(sheet: Worksheet, column: str) -> list[dict[str, Any]]:
    health = sheet.title == "Spending"
    label_cell = "H4" if health else "C3"
    if sheet["A3"].value != "$ millions" or sheet[label_cell].value != (
        "Health" if health else "Nominal GDP"
    ):
        raise ValueError("historical_headers")
    end = next(
        (
            row
            for row in range(5, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "% GDP"
        ),
        sheet.max_row + 1,
    )
    notes = _notes(sheet, end)
    period: dict[str, Any] = {}
    entries = []
    for row, year, markers, year_label in _year_rows(sheet, end):
        period = _period(sheet, row, period)
        if "period_end_month" not in period:
            raise ValueError("missing_historical_period")
        if any(marker not in notes for marker in markers):
            raise ValueError("missing_historical_footnote")
        sources = {
            "year": f"B{row}",
            "year_label": f"B{row}",
            "amount": f"{column}{row}",
            "period_end_month": period["period_cell"],
            "valid_time_end": period["period_cell"],
            "unit": "A3",
            "measure": label_cell,
        }
        if health:
            sources["accounting_basis"] = period["basis_cell"]
        entries.append(
            {
                **period,
                "year": year,
                "year_label": year_label,
                "sources": sources,
                "notes": [notes[marker] for marker in markers],
            }
        )
    return entries


def _amount_reason(cell: Cell | MergedCell, token: str | None) -> str | None:
    if cell.data_type == "f":
        return "formula_not_evaluated"
    if cell.data_type == "e":
        return "spreadsheet_error"
    if cell.value is None:
        return "missing_amount"
    if token is None:
        return "non_numeric_amount"
    if _exact_amount(token) is None:
        return "unsupported_precision"
    return None


def _fact(
    sheet: Worksheet, entry: dict[str, Any], token: str, context: dict[str, Any]
) -> dict[str, Any]:
    health = sheet.title == "Spending"
    sources: dict[str, str] = entry["sources"]
    cell = sheet[sources["amount"]]
    record_id = identity(
        _TRANSFORM, context["source_object_sha256"], sheet.title, cell.coordinate
    )
    period = entry["period_end_month"]
    return {
        **context,
        "record_id": record_id,
        "schema_version": "archive-govt-nz.health-historical-silver/v1",
        "domain": "health_appropriations",
        "recordset": "health_spending_fact" if health else "fiscal_context_fact",
        "year": entry["year"],
        "year_label": entry["year_label"],
        "amount": _exact_amount(token),
        "source_number_token": token,
        "source_number_format": cell.number_format,
        "period_end_month": period,
        "accounting_basis": entry["accounting_basis"],
        "valid_time_start": None,
        "valid_time_end": date(entry["year"], period, 31 if period == _MARCH else 30),
        "unit": "NZD_millions",
        "measure": "health_spending" if health else "nominal_gdp",
        "footnotes": [text for _, text in entry["notes"]],
        "rights_state": "not_evaluated",
        "quality_flags": [
            "period_start_not_provided",
            "cross_basis_comparability_not_asserted",
        ],
        "transformation_id": _TRANSFORM,
        "lineage_id": identity(record_id, "lineage"),
        "raw_values_json": encode_json(
            {
                coordinate: token if field == "amount" else sheet[coordinate].value
                for field, coordinate in sources.items()
            }
        ),
    }


def _rows(
    sheet: Worksheet, column: str, tokens: dict[str, str], context: dict[str, Any]
) -> tuple[
    list[dict[str, Any]], list[dict[str, Any]], dict[str, tuple[str, str, str | None]]
]:
    facts, lineage = [], []
    selected: dict[str, tuple[str, str, str | None]] = {}
    for entry in _inputs(sheet, column):
        sources: dict[str, str] = entry["sources"]
        uses: list[tuple[str, str]] = [
            *sources.items(),
            ("valid_time_end", sources["year"]),
            ("source_number_token", sources["amount"]),
            ("source_number_format", sources["amount"]),
            *(("footnotes", coordinate) for coordinate, _ in entry["notes"]),
        ]
        for _, coordinate in uses:
            selected[coordinate] = ("context", "historical_context", None)
        amount_cell = sources["amount"]
        token = tokens.get(amount_cell)
        reason = _amount_reason(sheet[amount_cell], token)
        if reason is not None:
            selected[amount_cell] = ("rejected", reason, None)
            continue
        fact = _fact(sheet, entry, str(token), context)
        facts.append(fact)
        selected[amount_cell] = (
            "normalized",
            "literal_historical_observation",
            fact["record_id"],
        )
        for field, coordinate in uses:
            lineage.append(
                {
                    "lineage_id": fact["lineage_id"],
                    "record_id": fact["record_id"],
                    "field": field,
                    "source_object_sha256": context["source_object_sha256"],
                    "source_locator": context["source_locator"],
                    "source_coordinate": _coordinate(sheet, coordinate),
                    "raw_value": token
                    if field in ("amount", "source_number_token")
                    else str(sheet[coordinate].number_format)
                    if field == "source_number_format"
                    else str(sheet[coordinate].value),
                    "normalized_value": str(fact[field]),
                    "rule": _TRANSFORM,
                }
            )
    return facts, lineage, selected


def normalize_historical_workbook(
    source: Path,
    output_dir: Path,
    *,
    expected_sha256: str,
    source_locator: str,
    source_vintage: str,
    observed_at: str,
) -> dict[str, object]:
    """Write separately identified historical currency observations from originals."""
    context = source_context(
        expected_sha256, source_locator, source_vintage, observed_at
    )
    payload = verified_snapshot(source, expected_sha256, max_bytes=64 * 1024 * 1024)
    inventory = inventory_workbook(BytesIO(payload))
    tokens = _number_tokens(payload)
    book = load_workbook(BytesIO(payload), data_only=False, keep_links=True)
    facts, lineage, dispositions = [], [], []
    try:
        for name, column in (("Spending", "H"), ("Nominal GDP", "C")):
            if name not in book.sheetnames:
                raise ValueError("missing_historical_sheet")
            sheet = book[name]
            sheet_facts, sheet_lineage, selected = _rows(
                sheet, column, tokens[name], context
            )
            facts.extend(sheet_facts)
            lineage.extend(sheet_lineage)
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None and cell.coordinate not in selected:
                        continue
                    disposition, reason, record_id = selected.get(
                        cell.coordinate,
                        (
                            "preserved_only",
                            "outside_historical_currency_selection",
                            None,
                        ),
                    )
                    dispositions.append(
                        {
                            "source_object_sha256": expected_sha256,
                            "source_coordinate": _coordinate(sheet, cell.coordinate),
                            "raw_value_json": encode_json(
                                tokens[name].get(cell.coordinate, cell.value)
                            ),
                            "disposition": disposition,
                            "reason": reason,
                            "record_id": record_id,
                        }
                    )
        excluded = [
            name for name in book.sheetnames if name not in ("Spending", "Nominal GDP")
        ]
    finally:
        book.close()
    rejected = sum(row["disposition"] == "rejected" for row in dispositions)
    return write_workbook_outputs(
        output_dir,
        {
            "historical_facts.parquet": pa.Table.from_pylist(facts, schema=_SCHEMA),
            "field_lineage.parquet": pa.Table.from_pylist(
                lineage, schema=LINEAGE_SCHEMA
            ),
            "cell_dispositions.parquet": pa.Table.from_pylist(
                dispositions, schema=_DISPOSITIONS
            ),
        },
        {
            "schema_version": "archive-govt-nz.health-historical-extraction/v1",
            "transformation_id": _TRANSFORM,
            "status": "partial" if rejected else "passed",
            "source_object_sha256": expected_sha256,
            "source_locator": source_locator,
            "source_vintage": source_vintage,
            "observed_at": context["observed_at"].isoformat(),
            "counts": {
                "facts": len(facts),
                "lineage": len(lineage),
                "dispositions": len(dispositions),
                "rejected": rejected,
            },
            "excluded_sheets": [
                {"sheet": name, "reason": "not_historical_health_gdp"}
                for name in excluded
            ],
            "rights_state": "not_evaluated",
            "workbook_inventory": inventory,
        },
    )
