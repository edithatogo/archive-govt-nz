"""Safe, non-mutating inventory of donor workbooks, PDFs, and SQLite."""

from __future__ import annotations

import re
import sqlite3
from dataclasses import asdict, dataclass
from typing import TYPE_CHECKING
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

_PDF_PAGE = re.compile(rb"/Type\s*/Page\b")
_MAX_MEMBERS = 20_000
_MAX_EXPANDED_BYTES = 512 * 1024 * 1024
_MAX_SCAN_CELLS = 2_000_000


@dataclass(frozen=True, slots=True)
class SheetInventory:
    """Structural workbook evidence for one sheet."""

    title: str
    state: str
    max_row: int
    max_column: int
    formula_cells: int
    merged_ranges: int
    table_names: tuple[str, ...]
    chart_count: int
    dimension: str
    formula_coordinates: tuple[str, ...]
    comment_coordinates: tuple[str, ...]
    merged_range_refs: tuple[str, ...]
    table_ranges: tuple[tuple[str, str], ...]
    hidden_rows: tuple[int, ...]
    hidden_columns: tuple[tuple[str, int | None, int | None], ...]
    defined_names: tuple[str, ...]


def _inventory_sheet(sheet: Worksheet) -> SheetInventory:
    formulas: list[str] = []
    comments: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                formulas.append(cell.coordinate)
            if cell.comment is not None:
                comments.append(cell.coordinate)
    merged = tuple(sorted(str(area) for area in sheet.merged_cells.ranges))
    return SheetInventory(
        title=sheet.title,
        state=sheet.sheet_state,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        formula_cells=len(formulas),
        merged_ranges=len(merged),
        table_names=tuple(sorted(sheet.tables)),
        chart_count=len(getattr(sheet, "_charts", ())),
        dimension=sheet.calculate_dimension(),
        formula_coordinates=tuple(formulas),
        comment_coordinates=tuple(comments),
        merged_range_refs=merged,
        table_ranges=tuple(
            (name, sheet.tables[name].ref) for name in sorted(sheet.tables)
        ),
        hidden_rows=tuple(
            sorted(index for index, dim in sheet.row_dimensions.items() if dim.hidden)
        ),
        hidden_columns=tuple(
            sorted(
                (name, dim.min, dim.max)
                for name, dim in sheet.column_dimensions.items()
                if dim.hidden
            )
        ),
        defined_names=tuple(sorted(sheet.defined_names)),
    )


def inventory_workbook(path: Path) -> dict[str, object]:
    """Inventory an XLSX package while leaving its original bytes untouched."""
    try:
        with ZipFile(path) as package:
            members = package.infolist()
            if len(members) > _MAX_MEMBERS:
                raise ValueError("workbook_member_limit")
            if any(
                member.filename.startswith("/")
                or "\\" in member.filename
                or ":" in member.filename
                or any(
                    part in ("", ".", "..")
                    for part in member.filename.removesuffix("/").split("/")
                )
                for member in members
            ):
                raise ValueError("unsafe_workbook_member")
            if len({member.filename for member in members}) != len(members):
                raise ValueError("duplicate_workbook_member")
            expanded = sum(member.file_size for member in members)
            if expanded > _MAX_EXPANDED_BYTES:
                raise ValueError("workbook_expansion_limit")
            package_members = tuple(sorted(member.filename for member in members))
    except BadZipFile:
        raise ValueError("invalid_workbook_package") from None

    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    sheets: list[SheetInventory] = []
    try:
        # Sparse sheets can have a huge rectangular extent despite tiny XML.
        # Bound the total traversal before iter_rows materializes empty cells.
        scan_cells = sum(
            sheet.max_row * sheet.max_column for sheet in workbook.worksheets
        )
        if scan_cells > _MAX_SCAN_CELLS:
            raise ValueError("workbook_cell_scan_limit")
        sheets = [_inventory_sheet(sheet) for sheet in workbook.worksheets]
        external_links = len(getattr(workbook, "_external_links", ()))
        defined_names = tuple(sorted(workbook.defined_names))
    finally:
        workbook.close()
    return {
        "schema_version": "archive-govt-nz.workbook-inventory/v1",
        "kind": "xlsx",
        "package_member_count": len(package_members),
        "package_members": package_members,
        "expanded_bytes": expanded,
        "has_macros": any(
            name.lower().endswith("vbaproject.bin") for name in package_members
        ),
        "external_link_count": external_links,
        "named_range_count": len(defined_names),
        "defined_names": defined_names,
        "sheets": [asdict(sheet) for sheet in sheets],
    }


def inventory_pdf(path: Path) -> dict[str, object]:
    """Return a bounded structural page count without rewriting the PDF."""
    payload = path.read_bytes()
    if not payload.startswith(b"%PDF-"):
        raise ValueError("invalid_pdf")
    return {
        "kind": "pdf",
        "byte_count": len(payload),
        "page_count": len(_PDF_PAGE.findall(payload)),
    }


def inventory_sqlite(path: Path) -> dict[str, object]:
    """Inventory user tables, schemas, and row counts read-only."""
    connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    try:
        tables = connection.execute(
            "SELECT name, sql FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
        return {
            "kind": "sqlite",
            "integrity": connection.execute("PRAGMA integrity_check").fetchone()[0],
            "tables": [
                {
                    "name": name,
                    "sql": sql,
                    "row_count": connection.execute(
                        f'SELECT COUNT(*) FROM "{name}"'
                    ).fetchone()[0],
                }
                for name, sql in tables
            ],
        }
    finally:
        connection.close()
