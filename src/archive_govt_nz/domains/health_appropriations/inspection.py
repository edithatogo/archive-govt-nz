"""Bounded, source-hash-bound workbook listings and decoded head previews."""

from __future__ import annotations

import json
from datetime import date, time, timedelta
from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook
from archive_govt_nz.domains.health_appropriations.workbook_common import (
    verified_snapshot,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

_MAX_SOURCE_BYTES = 64 * 1024 * 1024
_MAX_PREVIEW_CELLS = 2000
_MAX_PREVIEW_VALUE_BYTES = 128 * 1024
_MAX_ROWS = 20
_MAX_COLUMNS = 50


def _display(value: object) -> str | dict[str, Any]:
    if isinstance(value, (date, time, timedelta)):
        return str(value)
    if isinstance(value, (ArrayFormula, DataTableFormula)):
        result: dict[str, Any] = {"attributes": dict(value)}
        if isinstance(value, ArrayFormula):
            result["text"] = value.text
        return result
    message = "unsupported_preview_value"
    raise TypeError(message)


def _preview(
    sheet: Worksheet, rows: int, columns: int, remaining: int
) -> tuple[dict[str, Any], int]:
    cells = []
    for row in sheet.iter_rows(
        max_row=min(rows, sheet.max_row), max_col=min(columns, sheet.max_column)
    ):
        for cell in row:
            value = json.dumps(
                cell.value, ensure_ascii=False, default=_display, allow_nan=False
            )
            remaining -= len(value.encode("utf-8"))
            if remaining < 0:
                message = "preview_value_byte_limit"
                raise ValueError(message)
            cells.append(
                {
                    "coordinate": cell.coordinate,
                    "data_type": cell.data_type,
                    "decoded_value_json": value,
                }
            )
    return {
        "name": sheet.title,
        "row_truncated": sheet.max_row > rows,
        "column_truncated": sheet.max_column > columns,
        "cells": cells,
    }, remaining


def _inspect(
    source: Path, digest: str, sheet: str | None, rows: int, columns: int
) -> dict[str, Any]:
    if (
        type(rows) is not int
        or not 0 <= rows <= _MAX_ROWS
        or type(columns) is not int
        or not 1 <= columns <= _MAX_COLUMNS
    ):
        message = "invalid_preview_limits"
        raise ValueError(message)
    payload = verified_snapshot(source, digest, max_bytes=_MAX_SOURCE_BYTES)
    inventory = inventory_workbook(BytesIO(payload))
    workbook = load_workbook(BytesIO(payload), data_only=False, keep_links=True)
    try:
        sheets = {item.title: item for item in workbook.worksheets}
        if sheet is not None and sheet not in sheets:
            message = "unknown_worksheet"
            raise ValueError(message)
        selected = list(sheets.values()) if sheet is None else [sheets[sheet]]
        cells = sum(
            min(rows, item.max_row) * min(columns, item.max_column) for item in selected
        )
        if cells > _MAX_PREVIEW_CELLS:
            message = "preview_cell_limit"
            raise ValueError(message)
        previews = []
        remaining = _MAX_PREVIEW_VALUE_BYTES
        if rows:
            for item in selected:
                preview, remaining = _preview(item, rows, columns, remaining)
                previews.append(preview)
    finally:
        workbook.close()
    return {
        "schema_version": "archive-govt-nz.health-workbook-inspection/v1",
        "status": "inspected",
        "source_sha256": digest,
        "source_byte_count": len(payload),
        "inventory": inventory,
        "value_semantics": "decoded_preview_not_canonical_facts",
        "previews": previews,
    }


def inspect_workbook(
    source: Path,
    expected_sha256: str,
    *,
    sheet: str | None = None,
    rows: int = 5,
    columns: int = 12,
) -> dict[str, Any]:
    """Inspect one verified snapshot without evaluating formulas or writing files."""
    try:
        result = _inspect(source, expected_sha256, sheet, rows, columns)
    except Exception as error:  # noqa: BLE001 - redacted inspection boundary
        message = "workbook_inspection_failed:" + type(error).__name__
        raise ValueError(message) from None
    else:
        return result
