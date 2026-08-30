"""Budget expenditure facts from immutable, hash-verified original XLSX bytes.

No formula evaluation, donor-database dependency or redistribution approval is
implied. Outputs occupy a new directory; MANIFEST.json is the completion marker.
An interrupted write leaves an incomplete directory which must not be consumed.
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import TYPE_CHECKING, Any

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook
from archive_govt_nz.domains.health_appropriations.silver import (
    LINEAGE_SCHEMA,
    SILVER_SCHEMA,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.worksheet.worksheet import Worksheet

_MAX_SOURCE_BYTES = 64 * 1024 * 1024
_MAX_YEAR = 9999
_SHEET = "Raw Data"
_TRANSFORMATION = "budget-expenditure/v1"
_FIELDS = {
    "Year": "year",
    "Department": "department",
    "Appropriation Name": "appropriation_name",
    "Functional Classification": "functional_classification",
    "Amount $000": "amount",
    "Amount Type": "amount_type",
    "Portfolio Name": "portfolio_name",
}
_LABELS = tuple(name for name in _FIELDS if name not in ("Year", "Amount $000"))
_DISPOSITION_SCHEMA = pa.schema(
    [
        ("source_object_sha256", pa.string()),
        ("source_locator", pa.string()),
        ("sheet", pa.string()),
        ("source_row", pa.int64()),
        ("disposition", pa.string()),
        ("reason", pa.string()),
        ("record_id", pa.string()),
        ("raw_values_json", pa.string()),
    ]
)


def _identity(*parts: object) -> str:
    return "sha256:" + hashlib.sha256("\x1f".join(map(str, parts)).encode()).hexdigest()


def _json(value: object) -> str:
    return json.dumps(value, sort_keys=True, ensure_ascii=False, default=str)


def _number(value: object, *, year: bool = False) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = Decimal(str(value))
    except InvalidOperation:
        return None
    if not number.is_finite():
        return None
    if year:
        return number if 1 <= number <= _MAX_YEAR and number == int(number) else None
    # decimal128(20,3): 17 integral places, exactly representable; never round.
    if abs(number) >= Decimal("1e17"):
        return None
    scaled = number.quantize(Decimal("0.001"))
    return scaled if scaled == number else None


def _headers(sheet: Worksheet) -> list[str]:
    cells = next(sheet.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in cells]
    if (
        any(not isinstance(value, str) or not value.strip() for value in headers)
        or any(cell.data_type in ("f", "e") for cell in cells)
        or len(set(headers)) != len(headers)
        or not {"Vote", *_FIELDS}.issubset(headers)
    ):
        raise ValueError("invalid_headers")
    return [str(value) for value in headers]


def _classify(
    cells: tuple[Cell | MergedCell, ...], values: dict[str, Any]
) -> tuple[str, str]:
    if all(cell.value is None for cell in cells):
        return "blank", "empty_row"
    vote = values["Vote"]
    if (
        not isinstance(vote, str)
        or not vote.strip()
        or next(
            cell for cell, name in zip(cells, values, strict=True) if name == "Vote"
        ).data_type
        in ("f", "e")
    ):
        return "rejected", "invalid_vote"
    if vote != "Health":
        return "out_of_scope", "non_health_vote"
    if any(cell.data_type == "f" for cell in cells):
        return "rejected", "formula_not_evaluated"
    if any(cell.data_type == "e" for cell in cells):
        return "rejected", "spreadsheet_error"
    if any(
        not isinstance(values[name], str) or not values[name].strip()
        for name in _LABELS
    ):
        return "rejected", "missing_label"
    if _number(values["Year"], year=True) is None:
        return "rejected", "invalid_year"
    if _number(values["Amount $000"]) is None:
        return "rejected", "invalid_amount"
    return "normalized", "named_columns"


def _extract(
    sheet: Worksheet, context: dict[str, Any]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    headers = _headers(sheet)
    facts: list[dict[str, Any]] = []
    lineage: list[dict[str, Any]] = []
    dispositions: list[dict[str, Any]] = []
    for source_row, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = dict(zip(headers, (cell.value for cell in cells), strict=True))
        disposition, reason = _classify(cells, values)
        record_id = _identity(
            _TRANSFORMATION, context["source_object_sha256"], _SHEET, source_row
        )
        dispositions.append(
            {
                "source_object_sha256": context["source_object_sha256"],
                "source_locator": context["source_locator"],
                "sheet": _SHEET,
                "source_row": source_row,
                "disposition": disposition,
                "reason": reason,
                "record_id": record_id if disposition == "normalized" else None,
                "raw_values_json": _json(values),
            }
        )
        if disposition != "normalized":
            continue
        fact = {
            **context,
            "record_id": record_id,
            "schema_version": "archive-govt-nz.health-appropriations-silver/v1",
            "recordset": "appropriation_fact",
            "valid_time_start": None,
            "rights_state": "not_evaluated",
            "quality_flags": ["financial_year_basis_unverified"],
            "transformation_id": _TRANSFORMATION,
            "lineage_id": _identity(record_id, "lineage"),
            "donor_table": None,
            "donor_row_number": None,
            **{field: values[name] for name, field in _FIELDS.items()},
            "year": int(Decimal(str(values["Year"]))),
            "amount": _number(values["Amount $000"]),
            "measure": "appropriation_amount",
            "unit": "NZD_thousands",
            "raw_values_json": _json(values),
        }
        facts.append(fact)
        for name, cell in zip(headers, cells, strict=True):
            field = _FIELDS.get(name, f"raw:{name}")
            lineage.append(
                {
                    "lineage_id": fact["lineage_id"],
                    "record_id": record_id,
                    "field": field,
                    "source_object_sha256": context["source_object_sha256"],
                    "source_locator": context["source_locator"],
                    "source_coordinate": f"'{_SHEET}'!{cell.coordinate}",
                    "raw_value": str(cell.value),
                    "normalized_value": str(fact.get(field, cell.value)),
                    "rule": _TRANSFORMATION,
                }
            )
    return facts, lineage, dispositions


def normalize_budget_workbook(
    source: Path,
    output_dir: Path,
    *,
    expected_sha256: str,
    observed_at: str,
    source_vintage: str,
    source_locator: str,
) -> dict[str, object]:
    """Extract Health rows with dispositions and exact source-cell lineage.

    Source may be an extensionless Bronze CAS object. All parsing uses one
    verified in-memory snapshot. Other worksheets remain inventoried, not
    normalized. A passed receipt establishes extraction, not publication rights.
    """
    if re.fullmatch(r"[0-9a-f]{64}", expected_sha256) is None:
        raise ValueError("invalid_source_sha256")
    observed = datetime.fromisoformat(observed_at)
    if (
        observed.tzinfo is None
        or not source_vintage.strip()
        or not source_locator.strip()
    ):
        raise ValueError("invalid_source_context")
    with source.open("rb") as handle:
        payload = handle.read(_MAX_SOURCE_BYTES + 1)
    if len(payload) > _MAX_SOURCE_BYTES:
        raise ValueError("source_byte_limit")
    if hashlib.sha256(payload).hexdigest() != expected_sha256:
        raise ValueError("source_hash_mismatch")
    inventory = inventory_workbook(BytesIO(payload))
    workbook = load_workbook(BytesIO(payload), data_only=False, keep_links=True)
    try:
        if _SHEET not in workbook.sheetnames:
            raise ValueError("missing_raw_data_sheet")
        facts, lineage, dispositions = _extract(
            workbook[_SHEET],
            {
                "source_object_sha256": expected_sha256,
                "source_observation_id": _identity(
                    expected_sha256, source_locator, observed_at
                ),
                "source_locator": source_locator,
                "source_vintage": source_vintage,
                "observed_at": observed.astimezone(UTC),
            },
        )
        excluded = [
            {"sheet": name, "reason": "not_budget_raw_data"}
            for name in workbook.sheetnames
            if name != _SHEET
        ]
    finally:
        workbook.close()
    counts = {
        name: sum(row["disposition"] == name for row in dispositions)
        for name in ("normalized", "out_of_scope", "blank", "rejected")
    }
    counts["input"] = len(dispositions)
    outputs = {
        "budget_facts.parquet": pa.Table.from_pylist(facts, schema=SILVER_SCHEMA),
        "field_lineage.parquet": pa.Table.from_pylist(lineage, schema=LINEAGE_SCHEMA),
        "row_dispositions.parquet": pa.Table.from_pylist(
            dispositions, schema=_DISPOSITION_SCHEMA
        ),
    }
    # Reserve a new directory, never overwrite a prior result (even partial).
    output_dir.mkdir(parents=True, exist_ok=False)
    hashes = {}
    for name, table in outputs.items():
        path = output_dir / name
        pq.write_table(table, path)
        hashes[name] = hashlib.sha256(path.read_bytes()).hexdigest()
    receipt: dict[str, object] = {
        "schema_version": "archive-govt-nz.health-budget-extraction/v1",
        "transformation_id": _TRANSFORMATION,
        "status": "partial" if counts["rejected"] else "passed" if facts else "empty",
        "source_object_sha256": expected_sha256,
        "source_locator": source_locator,
        "source_vintage": source_vintage,
        "observed_at": observed.astimezone(UTC).isoformat(),
        "rights_state": "not_evaluated",
        "counts": counts,
        "excluded_sheets": excluded,
        "workbook_inventory": json.loads(_json(inventory)),
        "output_sha256": hashes,
    }
    # Last file is the completion marker; consumers must also verify its hashes.
    with (output_dir / "MANIFEST.json").open("x", encoding="utf-8") as handle:
        handle.write(_json(receipt) + "\n")
    return receipt
