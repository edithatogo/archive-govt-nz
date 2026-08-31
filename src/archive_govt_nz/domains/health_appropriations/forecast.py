"""Source-faithful BEFU/HYEFU literal Health expense summaries."""

from __future__ import annotations

import json
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from itertools import pairwise
from typing import TYPE_CHECKING, Any

import pyarrow as pa
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook
from archive_govt_nz.domains.health_appropriations.silver import (
    LINEAGE_SCHEMA,
    SILVER_SCHEMA,
)
from archive_govt_nz.domains.health_appropriations.workbook_common import (
    encode_json,
    exact_number,
    identity,
    source_context,
    verified_snapshot,
    write_workbook_outputs,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.worksheet.worksheet import Worksheet

_SHEETS = {
    "befu": "Core Crown Expense Tables",
    "hyefu": "Expense Tables",
    "befu-2026/v1": "Core Crown Expense Tables",
    "hyefu-2025/v1": "Core Crown Expense Tables",
}
FORECAST_PROFILES = tuple(_SHEETS)
_VINTAGE_ROWS = {"befu-2026/v1": ("BEFU-2026", 9), "hyefu-2025/v1": ("HYEFU-2025", 8)}
_VINTAGE_YEARS = tuple(str(year) for year in range(2021, 2031))
_VINTAGE_TYPES = ("Actual",) * 5 + ("Forecast",) * 5
_VINTAGE_COLUMNS = tuple(range(6, 16))
_MAX_SOURCE_BYTES = 64 * 1024 * 1024
_MIN_YEAR_COLUMNS = 2
_TRANSFORMATION = "treasury-health-expense-summary/v1"
TRANSFORMATION = _TRANSFORMATION
_CELL_SCHEMA = pa.schema(
    [
        ("source_object_sha256", pa.string()),
        ("source_locator", pa.string()),
        ("source_coordinate", pa.string()),
        ("data_type", pa.string()),
        ("raw_value_json", pa.string()),
        ("disposition", pa.string()),
        ("reason", pa.string()),
        ("record_id", pa.string()),
    ]
)


@dataclass(frozen=True, slots=True)
class _Layout:
    label: Cell | MergedCell
    unit: Cell | MergedCell
    columns: tuple[int, ...]
    year_row: int
    type_row: int
    amount_row: int


def _layout(sheet: Worksheet) -> _Layout:
    labels = [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if cell.value == "Health" and cell.data_type not in ("f", "e")
    ]
    if len(labels) != 1:
        raise ValueError("health_summary_label")
    label = labels[0]
    label_row, label_column = coordinate_to_tuple(label.coordinate)
    units = [
        sheet.cell(row, label_column)
        for row in range(1, label_row)
        if sheet.cell(row, label_column).value == "($millions)"
        and sheet.cell(row, label_column).data_type not in ("f", "e")
    ]
    if not units:
        raise ValueError("summary_unit")
    unit = units[-1]
    unit_row, _ = coordinate_to_tuple(unit.coordinate)
    if unit_row == 1:
        raise ValueError("summary_years")
    years = [
        sheet.cell(unit_row - 1, column)
        for column in range(label_column + 1, sheet.max_column + 1)
        if sheet.cell(unit_row - 1, column).value is not None
    ]
    if len(years) < _MIN_YEAR_COLUMNS or any(
        cell.data_type in ("f", "e") or exact_number(cell.value, year=True) is None
        for cell in years
    ):
        raise ValueError("summary_years")
    columns = tuple(coordinate_to_tuple(cell.coordinate)[1] for cell in years)
    if any(right != left + 1 for left, right in pairwise(columns)) or any(
        Decimal(str(right.value)) != Decimal(str(left.value)) + 1
        for left, right in pairwise(years)
    ):
        raise ValueError("summary_years")
    seen_forecast = False
    for column in columns:
        cell = sheet.cell(unit_row, column)
        if cell.data_type in ("f", "e") or cell.value not in ("Actual", "Forecast"):
            raise ValueError("summary_amount_types")
        if seen_forecast and cell.value == "Actual":
            raise ValueError("summary_amount_types")
        seen_forecast = seen_forecast or cell.value == "Forecast"
    if any(
        sheet.cell(label_row, column).value is not None
        for column in range(label_column + 1, sheet.max_column + 1)
        if column not in columns
    ):
        raise ValueError("unlabelled_summary_value")
    return _Layout(label, unit, columns, unit_row - 1, unit_row, label_row)


def _coordinate(sheet: Worksheet, cell: Cell | MergedCell) -> str:
    return f"'{sheet.title}'!{cell.coordinate}"


def _validate_vintage(
    sheet: Worksheet, layout: _Layout, profile: str, source_vintage: str
) -> None:
    """Bind reviewed successors explicitly; do not reinterpret legacy profiles."""
    if profile not in _VINTAGE_ROWS:
        return
    vintage, row = _VINTAGE_ROWS[profile]
    if (
        source_vintage != vintage
        or layout.label.coordinate != f"D{row}"
        or layout.unit.coordinate != f"D{row - 3}"
        or layout.columns != _VINTAGE_COLUMNS
        or tuple(sheet.cell(layout.year_row, column).value for column in layout.columns)
        != _VINTAGE_YEARS
        or tuple(sheet.cell(layout.type_row, column).value for column in layout.columns)
        != _VINTAGE_TYPES
    ):
        raise ValueError("forecast_vintage_contract")


def _amount_reason(cell: Cell | MergedCell) -> str | None:
    if cell.data_type == "f":
        return "formula_not_evaluated"
    if cell.data_type == "e":
        return "spreadsheet_error"
    if exact_number(cell.value) is None:
        return "invalid_amount"
    return None


def _extract(
    sheet: Worksheet, layout: _Layout, context: dict[str, Any]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    facts: list[dict[str, Any]] = []
    lineage: list[dict[str, Any]] = []
    selected: dict[str, tuple[str, str, str | None]] = {}
    for column in layout.columns:
        cells = {
            "year": sheet.cell(layout.year_row, column),
            "amount": sheet.cell(layout.amount_row, column),
            "amount_type": sheet.cell(layout.type_row, column),
            "functional_classification": layout.label,
            "measure": layout.label,
            "unit": layout.unit,
        }
        for field, cell in cells.items():
            if field != "amount":
                selected[cell.coordinate] = ("context", "summary_context", None)
        amount_cell = cells["amount"]
        reason = _amount_reason(amount_cell)
        if reason is not None:
            selected[amount_cell.coordinate] = ("rejected", reason, None)
            continue
        record_id = identity(
            _TRANSFORMATION,
            context["source_object_sha256"],
            sheet.title,
            amount_cell.coordinate,
        )
        selected[amount_cell.coordinate] = ("normalized", "literal_summary", record_id)
        fact = {
            **context,
            "record_id": record_id,
            "schema_version": "archive-govt-nz.health-appropriations-silver/v1",
            "recordset": "health_spending_fact",
            "valid_time_start": None,
            "rights_state": "not_evaluated",
            "quality_flags": ["financial_year_basis_unverified"],
            "transformation_id": _TRANSFORMATION,
            "lineage_id": identity(record_id, "lineage"),
            "donor_table": None,
            "donor_row_number": None,
            "year": int(Decimal(str(cells["year"].value))),
            "functional_classification": "Health",
            "amount_type": cells["amount_type"].value,
            "measure": "health_spending",
            "unit": "NZD_millions",
            "amount": exact_number(amount_cell.value),
            "raw_values_json": encode_json(
                {
                    cell.coordinate: {"value": cell.value, "data_type": cell.data_type}
                    for cell in cells.values()
                }
            ),
        }
        facts.append(fact)
        for field, cell in cells.items():
            lineage.append(
                {
                    "lineage_id": fact["lineage_id"],
                    "record_id": record_id,
                    "field": field,
                    "source_object_sha256": context["source_object_sha256"],
                    "source_locator": context["source_locator"],
                    "source_coordinate": _coordinate(sheet, cell),
                    "raw_value": str(cell.value),
                    "normalized_value": str(fact[field]),
                    "rule": _TRANSFORMATION,
                }
            )
    dispositions = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None and cell.coordinate not in selected:
                continue
            disposition, reason, record_id = selected.get(
                cell.coordinate,
                ("preserved_only", "outside_literal_health_summary", None),
            )
            dispositions.append(
                {
                    "source_object_sha256": context["source_object_sha256"],
                    "source_locator": context["source_locator"],
                    "source_coordinate": _coordinate(sheet, cell),
                    "data_type": cell.data_type,
                    "raw_value_json": encode_json(cell.value),
                    "disposition": disposition,
                    "reason": reason,
                    "record_id": record_id,
                }
            )
    return facts, lineage, dispositions


def normalize_forecast_workbook(
    source: Path,
    output_dir: Path,
    *,
    expected_sha256: str,
    profile: str,
    observed_at: str,
    source_vintage: str,
    source_locator: str,
    dry_run: bool = False,
) -> dict[str, object]:
    """Inspect or extract a verified summary without interpreting formula caches.

    Existing callers still write by default. Explicit ``dry_run=True`` reads
    and validates the same source but creates no Arrow tables or output state;
    rejected amounts remain partial, never a successful plan. This is not a
    serialization, filesystem-write or publication-readiness check.
    """
    if type(dry_run) is not bool:
        raise ValueError("forecast_dry_run_type")
    if profile not in _SHEETS:
        raise ValueError("unsupported_forecast_profile")
    context = source_context(
        expected_sha256, source_locator, source_vintage, observed_at
    )
    payload = verified_snapshot(source, expected_sha256, max_bytes=_MAX_SOURCE_BYTES)
    inventory = inventory_workbook(BytesIO(payload))
    workbook = load_workbook(BytesIO(payload), data_only=False, keep_links=True)
    try:
        if _SHEETS[profile] not in workbook.sheetnames:
            raise ValueError("missing_forecast_sheet")
        sheet = workbook[_SHEETS[profile]]
        layout = _layout(sheet)
        _validate_vintage(sheet, layout, profile, source_vintage)
        facts, lineage, dispositions = _extract(sheet, layout, context)
        excluded = [
            {"sheet": name, "reason": "not_forecast_summary_sheet"}
            for name in workbook.sheetnames
            if name != sheet.title
        ]
        selection = {
            "sheet": sheet.title,
            "label_cell": layout.label.coordinate,
            "unit_cell": layout.unit.coordinate,
            "year_row": layout.year_row,
            "amount_type_row": layout.type_row,
            "columns": list(layout.columns),
        }
    finally:
        workbook.close()
    counts = {
        name: sum(row["disposition"] == name for row in dispositions)
        for name in ("normalized", "rejected", "context", "preserved_only")
    }
    counts["inventoried_cells"] = len(dispositions)
    receipt = {
        "schema_version": "archive-govt-nz.health-forecast-extraction/v1",
        "transformation_id": _TRANSFORMATION,
        "profile": profile,
        "status": "partial"
        if counts["rejected"]
        else "planned"
        if dry_run
        else "passed",
        "source_object_sha256": expected_sha256,
        "source_locator": source_locator,
        "source_vintage": source_vintage,
        "observed_at": context["observed_at"].isoformat(),
        "rights_state": "not_evaluated",
        "counts": counts,
        "selection": selection,
        "disposition_scope": "nonempty_cells_plus_selected_inputs",
        "excluded_sheets": excluded,
        "workbook_inventory": inventory,
    }
    if dry_run:
        return {
            **json.loads(encode_json(receipt)),
            "preflight_scope": "source_validation_only",
        }
    outputs = {
        "forecast_facts.parquet": pa.Table.from_pylist(facts, schema=SILVER_SCHEMA),
        "field_lineage.parquet": pa.Table.from_pylist(lineage, schema=LINEAGE_SCHEMA),
        "cell_dispositions.parquet": pa.Table.from_pylist(
            dispositions, schema=_CELL_SCHEMA
        ),
    }
    return write_workbook_outputs(output_dir, outputs, receipt)
