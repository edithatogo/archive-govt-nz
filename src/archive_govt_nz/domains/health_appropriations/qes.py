"""Bounded published QES earnings, never an implicitly chosen deflator."""

from __future__ import annotations

import calendar
import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from types import MappingProxyType
from typing import TYPE_CHECKING, Any

import pyarrow as pa
from openpyxl import load_workbook

from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook
from archive_govt_nz.domains.health_appropriations.historical import _number_tokens
from archive_govt_nz.domains.health_appropriations.silver import LINEAGE_SCHEMA
from archive_govt_nz.domains.health_appropriations.workbook_common import (
    encode_json,
    identity,
    source_context,
    verified_snapshot,
    write_workbook_outputs,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

MAX_BYTES = 1024 * 1024
MAX_FIELD = 4096
TRANSFORMATION = "qes-june2026-table8-ordinary-hourly/v1"
RELEASE_TITLE = "Quarterly Employment Survey: June 2026 quarter"
HEADERS = MappingProxyType(
    {
        "A1": "Table 8",
        "A3": "Average hourly earnings(1)",
        "A4": "By sector",
        "P6": "Total",
        "P7": "Ordinary time",
        "A8": "Series ref: QEMQ",
        "P8": "SASZ9A",
        "A10": "($)",
        "A12": "Quarter",
        "A23": "Percentage change from the same quarter of previous year",
        "A36": "Percentage change from previous quarter",
        "B49": (
            "Average hourly earnings are calculated by dividing earnings by paid hours."
        ),
        "A51": "Source: Stats NZ",
    }
)
PERIODS = (
    (2024, 6),
    (2024, 9),
    (2024, 12),
    (2025, 3),
    (2025, 6),
    (2025, 9),
    (2025, 12),
    (2026, 3),
    (2026, 6),
)
FLAGS = (
    "currency_code_not_supplied",
    "sex_not_supplied",
    "adjustment_not_supplied",
    "not_selected_as_deflator",
    "published_not_independently_recomputed",
)
FACT_SCHEMA = pa.schema(
    [
        *(
            (key, pa.string())
            for key in (
                "record_id",
                "lineage_id",
                "transformation_id",
                "earnings_basis",
                "schema_version",
                "recordset",
                "source_object_sha256",
                "source_observation_id",
                "source_locator",
                "source_vintage",
                "series_id",
                "measure",
                "sector",
                "currency",
                "sex",
                "adjustment",
                "status",
                "unit_label",
                "unit_basis",
                "source_number_token",
                "source_number_format",
                "source_year_token",
                "source_quarter_token",
                "raw_values_json",
                "rights_state",
            )
        ),
        ("observed_at", pa.timestamp("us", tz="UTC")),
        ("amount", pa.decimal128(38, 18)),
        ("period_end", pa.date32()),
        ("quality_flags", pa.list_(pa.string())),
    ]
)
DISPOSITION_SCHEMA = pa.schema(
    [
        (key, pa.string())
        for key in (
            "source_object_sha256",
            "source_coordinate",
            "raw_value_json",
            "disposition",
            "reason",
            "record_id",
        )
    ]
)


def _number(token: str) -> Decimal:
    """Bound the exact source lexeme to Decimal128(38,18), without rounding."""
    if re.fullmatch(r"-?(?:0|[1-9][0-9]{0,19})(?:\.[0-9]{1,18})?", token) is None:
        message = "qes_invalid_literal_amount"
        raise ValueError(message)
    return Decimal(token)


def _extract(
    sheet: Worksheet, tokens: dict[str, str], context: dict[str, Any]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, str]]:
    if any(
        sheet[address].value != label or sheet[address].data_type in ("f", "e")
        for address, label in HEADERS.items()
    ):
        message = "qes_header_contract"
        raise ValueError(message)
    if sheet["P22"].value is not None:
        message = "qes_unlabelled_level"
        raise ValueError(message)
    facts, lineage, selected = [], [], {}
    for row, (year, month) in enumerate(PERIODS, 13):
        year_row = {2024: 13, 2025: 16, 2026: 20}[year]
        raw_year = sheet[f"A{row}"].value
        if (
            sheet[f"A{year_row}"].value != str(year)
            or (row != year_row and raw_year not in (None, ""))
            or sheet[f"C{row}"].value != calendar.month_abbr[month]
            or sheet[f"Q{row}"].value is not None
        ):
            message = "qes_period_or_status_contract"
            raise ValueError(message)
        address = f"P{row}"
        token = tokens.get(address, "")
        amount = _number(token)
        record = identity(
            TRANSFORMATION, context["source_object_sha256"], "Table 8", address
        )
        selected[address] = record
        fields = {
            "earnings_basis": "P7",
            "source_year_token": f"A{row}",
            "year_context": f"A{year_row}",
            "source_quarter_token": f"C{row}",
            "amount": address,
            **{f"header_{key}": key for key in HEADERS},
        }
        raw = {key: sheet[cell].value for key, cell in fields.items()}
        raw["amount"] = token
        facts.append(
            {
                **context,
                "record_id": record,
                "lineage_id": identity(record, "lineage"),
                "transformation_id": TRANSFORMATION,
                "earnings_basis": "ordinary_time",
                "schema_version": "archive-govt-nz.qes-earnings/v1",
                "recordset": "published_earnings_fact",
                "series_id": "QEMQ.SASZ9A",
                "measure": "Average hourly earnings",
                "sector": "Total",
                "currency": None,
                "sex": None,
                "adjustment": None,
                "status": None,
                "unit_label": "($)",
                "unit_basis": "per_paid_hour",
                "amount": amount,
                "source_number_token": token,
                "source_number_format": sheet[address].number_format,
                "source_year_token": raw_year,
                "source_quarter_token": sheet[f"C{row}"].value,
                "period_end": date(year, month, calendar.monthrange(year, month)[1]),
                "raw_values_json": encode_json(raw),
                "rights_state": "not_evaluated",
                "quality_flags": list(FLAGS),
            }
        )
        for field, coordinate in fields.items():
            lineage.append(
                {
                    "lineage_id": identity(record, field),
                    "record_id": record,
                    "field": field,
                    "source_object_sha256": context["source_object_sha256"],
                    "source_locator": context["source_locator"],
                    "source_coordinate": f"'Table 8'!{coordinate}",
                    "raw_value": encode_json(raw[field]),
                    "normalized_value": encode_json(
                        "ordinary_time" if field == "earnings_basis" else raw[field]
                    ),
                    "rule": "literal_qes_source_context",
                }
            )
        lineage.extend(
            {
                "lineage_id": identity(record, "period_end", coordinate),
                "record_id": record,
                "field": "period_end",
                "source_object_sha256": context["source_object_sha256"],
                "source_locator": context["source_locator"],
                "source_coordinate": f"'Table 8'!{coordinate}",
                "raw_value": encode_json(sheet[coordinate].value),
                "normalized_value": encode_json(
                    date(year, month, calendar.monthrange(year, month)[1]).isoformat()
                ),
                "rule": "validated_calendar_quarter_end",
            }
            for coordinate in (f"A{year_row}", f"C{row}")
        )
    return facts, lineage, selected


def _dispositions(
    sheet: Worksheet,
    tokens: dict[str, dict[str, str]],
    selected: dict[str, str],
    sha256: str,
) -> list[dict[str, Any]]:
    dispositions = []
    for row in sheet:
        for cell in row:
            if cell.value is None:
                continue
            record = selected.get(cell.coordinate) if sheet.title == "Table 8" else None
            reason = "literal_earnings" if record else "outside_selected_series"
            if sheet.title == "Table 8" and cell.row in range(23, 48):
                reason = "published_percentage_change_not_level"
            dispositions.append(
                {
                    "source_object_sha256": sha256,
                    "source_coordinate": f"'{sheet.title}'!{cell.coordinate}",
                    "raw_value_json": encode_json(
                        tokens[sheet.title].get(cell.coordinate, cell.value)
                    ),
                    "disposition": "normalized" if record else "preserved_only",
                    "reason": reason,
                    "record_id": record,
                }
            )
    return dispositions


def normalize_qes(  # noqa: PLR0913 - explicit source provenance is required.
    source: Path,
    output_dir: Path,
    *,
    expected_sha256: str,
    source_vintage: str,
    source_locator: str,
    observed_at: str,
    dry_run: bool = True,
) -> dict[str, Any]:
    """Read a reviewed nine-quarter profile; new local outputs require opt-in.

    The historical reader's internal literal-token utility is deliberately reused
    unchanged after the same workbook inventory admission; this is not a generic
    hostile-workbook sandbox. No source acquisition or rights assessment occurs.
    """
    if source.is_symlink() or output_dir.is_symlink():
        message = "qes_symlink_path"
        raise ValueError(message)
    if source_vintage != "QES-2026-Q2":
        message = "qes_vintage_contract"
        raise ValueError(message)
    context = source_context(
        expected_sha256, source_locator, source_vintage, observed_at
    )
    payload = verified_snapshot(source, expected_sha256, max_bytes=MAX_BYTES)
    inventory = inventory_workbook(BytesIO(payload))
    tokens = _number_tokens(payload)
    workbook = load_workbook(BytesIO(payload), data_only=False)
    try:
        if "Table 8" not in workbook.sheetnames:
            message = "qes_missing_sheet"
            raise ValueError(message)
        if (
            "Contents" not in workbook.sheetnames
            or workbook["Contents"]["A1"].value != RELEASE_TITLE
        ):
            message = "qes_release_contract"
            raise ValueError(message)
        for sheet in workbook:
            for row in sheet:
                if any(len(str(cell.value)) > MAX_FIELD for cell in row):
                    message = "qes_field_limit"
                    raise ValueError(message)
        facts, lineage, selected = _extract(
            workbook["Table 8"], tokens["Table 8"], context
        )
        dispositions = [
            entry
            for sheet in workbook
            for entry in _dispositions(sheet, tokens, selected, expected_sha256)
        ]
    finally:
        workbook.close()
    receipt = {
        "schema_version": "archive-govt-nz.qes-extraction/v1",
        "transformation_id": TRANSFORMATION,
        "status": "passed",
        "source_object_sha256": expected_sha256,
        "source_locator": source_locator,
        "source_vintage": source_vintage,
        "observed_at": context["observed_at"].isoformat(),
        "rights_state": "not_evaluated",
        "counts": {
            "normalized": len(facts),
            "field_lineage": len(lineage),
            "inventoried_cells": len(dispositions),
        },
        "workbook_inventory": inventory,
        "disposition_scope": "nonempty_cells",
        "quality_flags": list(FLAGS),
    }
    if dry_run:
        return receipt
    return write_workbook_outputs(
        output_dir,
        {
            "qes_facts.parquet": pa.Table.from_pylist(facts, schema=FACT_SCHEMA),
            "field_lineage.parquet": pa.Table.from_pylist(
                lineage, schema=LINEAGE_SCHEMA
            ),
            "cell_dispositions.parquet": pa.Table.from_pylist(
                dispositions, schema=DISPOSITION_SCHEMA
            ),
        },
        receipt,
    )
