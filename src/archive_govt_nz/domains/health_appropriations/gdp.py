"""Literal quarterly GDP expenditure actuals, without denominator selection."""

from __future__ import annotations

import calendar
import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Any

import pyarrow as pa
from openpyxl import load_workbook

from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook
from archive_govt_nz.domains.health_appropriations.historical import _number_tokens
from archive_govt_nz.domains.health_appropriations.silver import LINEAGE_SCHEMA
from archive_govt_nz.domains.health_appropriations.workbook_common import (
    encode_json,
    identity,
    source_context,
    verified_snapshot,
    write_workbook_outputs,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook

MAX_BYTES = 1024 * 1024
VINTAGE = "StatsNZ-GDP-2026Q1"
TRANSFORMATION = "stats-nz-gdp-current-price-expenditure-actual-2026q1/v1"
TITLE = (
    "Gross domestic product: March 2026 quarter "
    "\u2013 current price income and expenditure"
)
HEADERS = {
    "A2": "Gross domestic product – current price income and expenditure(1)",  # noqa: RUF001 - exact publisher label
    "A4": "Quarterly, June 2011–March 2026",  # noqa: RUF001 - exact publisher label
    "B5": "Series",
    "B6": "ref:",
    "B7": "SNEQ",
    "C5": "Quarter",
    "C7": "$(million)",
    "A27": "Gross domestic product - expenditure measure",
    "B27": "SG03AB01GE00S900",
    "A30": "1. Data may not sum to totals due to rounding.",
    "A31": "Source: Stats NZ",
}
_MONTH_NAMES = {3: "Mar", 6: "Jun", 9: "Sep", 12: "Dec"}
PERIODS = tuple(
    (
        f"{_MONTH_NAMES[month]}-{year % 100:02d}",
        date(year, month, calendar.monthrange(year, month)[1]),
    )
    for year, month in (
        (2011 + (quarter + 1) // 4, ((quarter + 1) % 4 + 1) * 3)
        for quarter in range(60)
    )
)
FACT_SCHEMA = pa.schema(
    [
        (key, pa.string())
        for key in (
            "record_id",
            "schema_version",
            "recordset",
            "source_object_sha256",
            "source_observation_id",
            "source_locator",
            "source_vintage",
            "source_coordinate",
            "series_prefix",
            "series_reference",
            "label",
            "period_token",
            "source_number_token",
            "source_number_format",
            "unit",
            "scaling",
            "currency",
            "price_basis",
            "adjustment",
            "footnote",
            "rights_state",
            "transformation_id",
            "lineage_id",
            "raw_values_json",
        )
    ]
    + [
        ("observed_at", pa.timestamp("us", tz="UTC")),
        ("period_end", pa.date32()),
        ("publication_date", pa.date32()),
        ("amount", pa.decimal128(20, 0)),
        ("quality_flags", pa.list_(pa.string())),
    ]
)
DISPOSITION_SCHEMA = pa.schema(
    [
        (key, pa.string())
        for key in (
            "source_object_sha256",
            "source_coordinate",
            "raw_value_json",
            "disposition",
            "reason",
            "record_id",
        )
    ]
)


def _require(condition: object) -> None:
    if not condition:
        message = "gdp_source_contract"
        raise ValueError(message)


def _amount(token: str) -> Decimal:
    _require(re.fullmatch(r"-?[0-9]{1,20}", token) is not None)
    return Decimal(token)


def _profile(book: Workbook) -> None:
    _require(book.sheetnames == ["Contents", "Table 1", "Table 2"])
    sheet = book["Table 1"]
    _require((sheet.max_row, sheet.max_column) == (31, 62))
    _require(all(sheet[key].value == value for key, value in HEADERS.items()))
    _require(sheet["A1"].value == "Table 1")
    _require(sheet["A3"].value == "Actual current prices")
    _require(
        {str(item) for item in sheet.merged_cells.ranges}
        == {
            "C5:BJ5",
            "C7:BJ7",
        }
    )
    _require(book["Contents"]["A1"].value == TITLE)
    _require(book["Contents"]["A33"].value == "18 June 2026")
    _require(book["Table 2"]["A3"].value == "Seasonally adjusted current prices")


def _extract(
    book: Workbook,
    tokens: dict[str, dict[str, str]],
    context: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    _profile(book)
    sheet = book["Table 1"]
    facts, lineage, dispositions = [], [], []
    selected, used = {}, set()
    for column, (period_token, period_end) in enumerate(PERIODS, 3):
        cell, header = sheet.cell(27, column), sheet.cell(6, column)
        token = tokens["Table 1"].get(cell.coordinate, "")
        _require(header.value == period_token and cell.data_type == "n")
        amount = _amount(token)
        _require(cell.number_format == "#,##0")
        coordinate = f"'Table 1'!{cell.coordinate}"
        record_id = identity(
            TRANSFORMATION,
            context["source_object_sha256"],
            coordinate,
            HEADERS["B27"],
            period_token,
        )
        selected[coordinate] = record_id
        raw = {
            "amount": (coordinate, token),
            "source_number_token": (coordinate, token),
            "source_number_format": (f"{coordinate}@number_format", cell.number_format),
            "period_token": (f"'Table 1'!{header.coordinate}", period_token),
            "period_end": (f"'Table 1'!{header.coordinate}", period_token),
            "period_end:range": ("'Table 1'!A4", HEADERS["A4"]),
            "series_prefix": ("'Table 1'!B7", HEADERS["B7"]),
            "series_reference": ("'Table 1'!B27", HEADERS["B27"]),
            "label": ("'Table 1'!A27", HEADERS["A27"]),
            "unit": ("'Table 1'!C7", HEADERS["C7"]),
            "scaling": ("'Table 1'!C7", HEADERS["C7"]),
            "price_basis": ("'Table 1'!A3", "Actual current prices"),
            "adjustment": ("'Table 1'!A3", "Actual current prices"),
            "footnote": ("'Table 1'!A30", HEADERS["A30"]),
            "publication_date": ("'Contents'!A33", "18 June 2026"),
        }
        fact = {
            **context,
            "record_id": record_id,
            "schema_version": "archive-govt-nz.health-gdp-silver/v1",
            "recordset": "economic_context_fact",
            "source_coordinate": coordinate,
            "series_prefix": HEADERS["B7"],
            "series_reference": HEADERS["B27"],
            "label": HEADERS["A27"],
            "period_token": period_token,
            "period_end": period_end,
            "amount": amount,
            "source_number_token": token,
            "source_number_format": cell.number_format,
            "unit": HEADERS["C7"],
            "scaling": "million",
            "currency": None,
            "price_basis": "current_prices",
            "adjustment": "actual_as_published",
            "footnote": HEADERS["A30"],
            "publication_date": date(2026, 6, 18),
            "rights_state": "not_evaluated",
            "transformation_id": TRANSFORMATION,
            "lineage_id": identity(record_id, "lineage"),
            "raw_values_json": encode_json(raw),
            "quality_flags": [
                "iso_currency_not_explicit",
                "published_rounding_retained",
                "quarterly_not_annual",
                "denominator_not_selected",
            ],
        }
        facts.append(fact)
        for raw_field, (source_coordinate, value) in raw.items():
            used.add(source_coordinate)
            field = raw_field.split(":", 1)[0]
            lineage.append(
                {
                    "lineage_id": fact["lineage_id"],
                    "record_id": record_id,
                    "field": field,
                    "source_object_sha256": context["source_object_sha256"],
                    "source_locator": context["source_locator"],
                    "source_coordinate": source_coordinate,
                    "raw_value": value,
                    "normalized_value": str(fact[field]),
                    "rule": TRANSFORMATION,
                }
            )
    for source_sheet in book:
        for row in source_sheet:
            for cell in row:
                if cell.value is None:
                    continue
                coordinate = f"'{source_sheet.title}'!{cell.coordinate}"
                disposition = "preserved_only"
                if coordinate in used:
                    disposition = "context"
                if coordinate in selected:
                    disposition = "selected"
                dispositions.append(
                    {
                        "source_object_sha256": context["source_object_sha256"],
                        "source_coordinate": coordinate,
                        "raw_value_json": encode_json(
                            tokens[source_sheet.title].get(cell.coordinate, cell.value)
                        ),
                        "disposition": disposition,
                        "reason": {
                            "selected": "exact_expenditure_actual_series",
                            "context": "selected_field_context",
                            "preserved_only": "outside_selected_series",
                        }[disposition],
                        "record_id": selected.get(coordinate),
                    }
                )
    return facts, lineage, dispositions


def normalize_gdp(  # noqa: PLR0913 - explicit provenance and safe dry-run contract
    source: Path,
    output_dir: Path,
    *,
    expected_sha256: str,
    source_locator: str,
    source_vintage: str,
    observed_at: str,
    dry_run: bool = True,
) -> dict[str, object]:
    """Extract only reviewed quarterly actuals; never aggregate or select ratios."""
    context = source_context(
        expected_sha256, source_locator, source_vintage, observed_at
    )
    _require(source_vintage == VINTAGE)
    _require(source.is_file() and not source.is_symlink())
    _require(not output_dir.exists() and not output_dir.is_symlink())
    payload = verified_snapshot(source, expected_sha256, max_bytes=MAX_BYTES)
    inventory_workbook(BytesIO(payload))
    tokens = _number_tokens(payload)
    book = load_workbook(BytesIO(payload), data_only=False, keep_links=True)
    try:
        facts, lineage, dispositions = _extract(book, tokens, context)
    finally:
        book.close()
    receipt = {
        "schema_version": "archive-govt-nz.health-gdp-extraction/v1",
        "transformation_id": TRANSFORMATION,
        "status": "planned" if dry_run else "passed",
        **context,
        "observed_at": context["observed_at"].isoformat(),
        "rights_state": "not_evaluated",
        "currency": None,
        "counts": {
            "facts": len(facts),
            "lineage": len(lineage),
            "dispositions": len(dispositions),
        },
    }
    if dry_run:
        return receipt
    return write_workbook_outputs(
        output_dir,
        {
            "gdp_facts.parquet": pa.Table.from_pylist(facts, schema=FACT_SCHEMA),
            "field_lineage.parquet": pa.Table.from_pylist(
                lineage, schema=LINEAGE_SCHEMA
            ),
            "cell_dispositions.parquet": pa.Table.from_pylist(
                dispositions, schema=DISPOSITION_SCHEMA
            ),
        },
        receipt,
    )
