"""Synthetic exact-profile tests; no official payload fixtures."""

from __future__ import annotations

import hashlib
import json
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

import pyarrow.parquet as pq
import pytest
from hypothesis import given
from hypothesis import strategies as st
from openpyxl import Workbook, load_workbook

from archive_govt_nz.domains.health_appropriations import qes
from archive_govt_nz.domains.health_appropriations.qes import _number


def fixture(path: Path, change: tuple[str, Any] | None = None) -> str:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Table 8"
    cells = {
        "A1": "Table 8",
        "A3": "Average hourly earnings(1)",
        "A4": "By sector",
        "P6": "Total",
        "P7": "Ordinary time",
        "A8": "Series ref: QEMQ",
        "P8": "SASZ9A",
        "A10": "($)",
        "A12": "Quarter",
        "A23": "Percentage change from the same quarter of previous year",
        "A36": "Percentage change from previous quarter",
        "B49": "Average hourly earnings are calculated by dividing earnings by paid hours.",
        "A51": "Source: Stats NZ",
        "A13": "2024",
        "A16": "2025",
        "A20": "2026",
        "P26": 1.5,
        "P39": 0.5,
    }
    for address, value in cells.items():
        sheet[address] = value
    for row, month in enumerate(
        ("Jun", "Sep", "Dec", "Mar", "Jun", "Sep", "Dec", "Mar", "Jun"), 13
    ):
        sheet[f"C{row}"] = month
        sheet[f"P{row}"] = row + 0.25
    if change:
        sheet[change[0]] = change[1]
    workbook.create_sheet("Table 9")["P13"] = 13.25
    workbook.create_sheet("Contents")["A1"] = (
        "Quarterly Employment Survey: June 2026 quarter"
    )
    workbook.save(path)
    workbook.close()
    return hashlib.sha256(path.read_bytes()).hexdigest()


def run(
    source: Path, output: Path, pin: str, *, dry_run: bool = True
) -> dict[str, Any]:
    return qes.normalize_qes(
        source,
        output,
        expected_sha256=pin,
        source_vintage="QES-2026-Q2",
        source_locator="https://example.invalid/qes.xlsx",
        observed_at="2026-08-31T00:00:00Z",
        dry_run=dry_run,
    )


def test_exact_profile_and_dry_run(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    before = source.read_bytes()
    output = tmp_path / "out"
    assert run(source, output, pin)["counts"]["normalized"] == 9
    assert not output.exists()
    receipt = run(source, output, pin, dry_run=False)
    facts = pq.read_table(output / "qes_facts.parquet").to_pylist()
    assert len(facts) == 9
    assert facts[0]["amount"] == Decimal("13.25")
    assert facts[0]["period_end"] == date(2024, 6, 30)
    assert facts[-1]["period_end"] == date(2026, 6, 30)
    assert all(
        f["currency"] is None and f["adjustment"] is None and f["sex"] is None
        for f in facts
    )
    assert all(f["series_id"] == "QEMQ.SASZ9A" for f in facts)
    assert receipt["rights_state"] == "not_evaluated"
    assert source.read_bytes() == before
    lineage = pq.read_table(output / "field_lineage.parquet").to_pylist()
    dispositions = pq.read_table(output / "cell_dispositions.parquet").to_pylist()
    assert len(lineage) == 9 * 20
    assert len({line["lineage_id"] for line in lineage}) == len(lineage)
    for row, fact in enumerate(facts, 13):
        assert fact["amount"] == Decimal(row) + Decimal("0.25")
        assert fact["source_number_token"] == f"{row}.25"
        assert fact["recordset"] == "published_earnings_fact"
        assert fact["quality_flags"] == list(qes.FLAGS)
        lines = [
            line
            for line in lineage
            if line["record_id"] == fact["record_id"] and line["field"] != "period_end"
        ]
        assert len(lines) == 18
        assert {
            line["field"]: json.loads(line["raw_value"]) for line in lines
        } == json.loads(fact["raw_values_json"])
        matches = [
            item for item in dispositions if item["record_id"] == fact["record_id"]
        ]
        assert len(matches) == 1
        assert matches[0]["source_coordinate"] == f"'Table 8'!P{row}"
        assert json.loads(matches[0]["raw_value_json"]) == f"{row}.25"
    preserved = {item["source_coordinate"]: item for item in dispositions}
    assert preserved["'Table 9'!P13"]["disposition"] == "preserved_only"
    assert (
        preserved["'Table 8'!P26"]["reason"] == "published_percentage_change_not_level"
    )
    assert (
        preserved["'Table 8'!P39"]["reason"] == "published_percentage_change_not_level"
    )
    for name, digest in receipt["output_sha256"].items():
        assert hashlib.sha256((output / name).read_bytes()).hexdigest() == digest
    second = tmp_path / "second"
    run(source, second, pin, dry_run=False)
    assert {p.name: p.read_bytes() for p in output.iterdir()} == {
        p.name: p.read_bytes() for p in second.iterdir()
    }


@pytest.mark.parametrize(
    "change",
    [
        ("P8", "OTHER"),
        ("A13", "2023"),
        ("C14", "Jun"),
        ("P13", "=1+2"),
        ("P13", "unknown"),
        ("P22", 7),
        ("A14", "2024"),
        ("Q13", "R"),
    ],
)
def test_contract_rejected_before_output(
    tmp_path: Path, change: tuple[str, object]
) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source, change)
    with pytest.raises(ValueError, match="qes_"):
        run(source, tmp_path / "out", pin, dry_run=False)
    assert not (tmp_path / "out").exists()


@pytest.mark.parametrize(
    "token",
    [
        "",
        "NaN",
        "Infinity",
        "1e2",
        " 1",
        "1 ",
        "+1",
        "01",
        "1.0000000000000000001",
        "100000000000000000000",
    ],
)
def test_reject_inexact_or_ambiguous_number(token: str) -> None:
    with pytest.raises(ValueError, match="qes_invalid_literal_amount"):
        _number(token)


@given(st.integers(min_value=-99999999, max_value=99999999))
def test_decimal_property(value: int) -> None:
    token = f"{value}.123456789012345678"
    assert _number(token) == Decimal(token)


def test_exact_decimal_boundary() -> None:
    number = "99999999999999999999.999999999999999999"
    assert _number(number) == Decimal(number)


@pytest.mark.parametrize("delta", [-1, 0, 1])
def test_source_cap(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, delta: int
) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    monkeypatch.setattr(qes, "MAX_BYTES", source.stat().st_size + delta)
    if delta < 0:
        with pytest.raises(ValueError, match="source_byte_limit"):
            run(source, tmp_path / "out", pin)
    else:
        assert run(source, tmp_path / "out", pin)["status"] == "passed"


@pytest.mark.parametrize("size", [4095, 4096, 4097])
def test_field_cap(tmp_path: Path, size: int) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source, ("Z1", "x" * size))
    if size > qes.MAX_FIELD:
        with pytest.raises(ValueError, match="qes_field_limit"):
            run(source, tmp_path / "out", pin)
    else:
        assert run(source, tmp_path / "out", pin)["status"] == "passed"


@pytest.mark.parametrize("target", ["source", "output", "dangling"])
def test_symlinks(tmp_path: Path, target: str) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    output = tmp_path / "out"
    if target == "source":
        link = tmp_path / "link"
        link.symlink_to(source)
        source = link
    else:
        directory = tmp_path / "directory"
        if target == "output":
            directory.mkdir()
        output.symlink_to(directory, target_is_directory=True)
    with pytest.raises(ValueError, match="qes_symlink_path"):
        run(source, output, pin, dry_run=False)


def test_exclusive_and_failed_partial(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    output = tmp_path / "out"
    output.mkdir()
    with pytest.raises(FileExistsError):
        run(source, output, pin, dry_run=False)
    assert list(output.iterdir()) == []

    def fail(*_args: object, **_kwargs: object) -> None:
        message = "synthetic interruption"
        raise OSError(message)

    monkeypatch.setattr(pq, "write_table", fail)
    partial = tmp_path / "partial"
    with pytest.raises(OSError, match="synthetic interruption"):
        run(source, partial, pin, dry_run=False)
    assert (partial / "qes_facts.parquet").exists()
    assert not (partial / "MANIFEST.json").exists()


def test_missing_sheet_and_vintage(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    with pytest.raises(ValueError, match="qes_vintage_contract"):
        qes.normalize_qes(
            source,
            tmp_path / "out",
            expected_sha256=pin,
            source_vintage="wrong",
            source_locator="x",
            observed_at="2026-08-31T00:00:00Z",
        )
    book = Workbook()
    book.save(source)
    book.close()
    pin = hashlib.sha256(source.read_bytes()).hexdigest()
    with pytest.raises(ValueError, match="qes_missing_sheet"):
        run(source, tmp_path / "out", pin)


def test_release_title_drift(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    fixture(source)
    book = load_workbook(source)
    book["Contents"]["A1"] = "Quarterly Employment Survey: June 2027 quarter"
    book.save(source)
    book.close()
    pin = hashlib.sha256(source.read_bytes()).hexdigest()
    with pytest.raises(ValueError, match="qes_release_contract"):
        run(source, tmp_path / "out", pin)


def test_wrong_pin_and_context(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    with pytest.raises(ValueError, match="source_hash_mismatch"):
        run(source, tmp_path / "out", "0" * 64)
    with pytest.raises(ValueError, match="invalid_source_context"):
        qes.normalize_qes(
            source,
            tmp_path / "out",
            expected_sha256=pin,
            source_vintage="QES-2026-Q2",
            source_locator="",
            observed_at="2026-08-31T00:00:00Z",
        )


def test_explicit_basis_and_derived_period_lineage(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    pin = fixture(source)
    output = tmp_path / "out"
    run(source, output, pin, dry_run=False)
    facts = pq.read_table(output / "qes_facts.parquet").to_pylist()
    lines = pq.read_table(output / "field_lineage.parquet").to_pylist()
    for fact in facts:
        assert fact["earnings_basis"] == "ordinary_time"
        assert fact["transformation_id"] == qes.TRANSFORMATION
        assert fact["lineage_id"].startswith("sha256:")
        owned = [line for line in lines if line["record_id"] == fact["record_id"]]
        basis = [line for line in owned if line["field"] == "earnings_basis"]
        assert len(basis) == 1
        assert basis[0]["source_coordinate"] == "'Table 8'!P7"
        assert json.loads(basis[0]["normalized_value"]) == "ordinary_time"
        period = [line for line in owned if line["field"] == "period_end"]
        assert len(period) == 2
        assert len({line["lineage_id"] for line in period}) == 2
        assert all(
            json.loads(line["normalized_value"]) == fact["period_end"].isoformat()
            for line in period
        )
        assert all(line["rule"] == "validated_calendar_quarter_end" for line in period)
