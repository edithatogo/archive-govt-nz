"""Synthetic strict quarterly GDP profile and preservation contracts."""

from __future__ import annotations

import calendar
import hashlib
import importlib
import json
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

import pyarrow.parquet as pq
import pytest
from hypothesis import given
from hypothesis import strategies as st
from openpyxl import Workbook, load_workbook

from archive_govt_nz.domains.health_appropriations import gdp, workbook_common


def workbook(path: Path) -> Path:
    book = Workbook()
    assert book.active is not None
    book.active.title = "Contents"
    for name in ("Table 1", "Table 2"):
        sheet = book.create_sheet(name)
        for coordinate, value in gdp.HEADERS.items():
            sheet[coordinate] = value
        sheet["A1"] = name
        sheet["A3"] = (
            "Actual current prices"
            if name == "Table 1"
            else "Seasonally adjusted current prices"
        )
        sheet.merge_cells("C5:BJ5")
        sheet.merge_cells("C7:BJ7")
        for column, (token, _) in enumerate(gdp.PERIODS, 3):
            sheet.cell(6, column, token)
            sheet.cell(27, column, 50000 + column).number_format = "#,##0"
    book["Contents"]["A1"] = gdp.TITLE
    book["Contents"]["A33"] = "18 June 2026"
    book.save(path)
    book.close()
    return path


def test_dry_run_preserves_original(tmp_path: Path) -> None:
    source = workbook(tmp_path / "source.xlsx")
    before = source.read_bytes()
    output = tmp_path / "absent" / "output"
    receipt = gdp.normalize_gdp(
        source,
        output,
        expected_sha256=hashlib.sha256(before).hexdigest(),
        source_locator="https://example.invalid/gdp.xlsx",
        source_vintage=gdp.VINTAGE,
        observed_at="2026-08-29T09:00:17Z",
    )
    counts = receipt["counts"]
    assert isinstance(counts, dict)
    assert counts["facts"] == 60
    assert receipt["currency"] is None
    assert not output.parent.exists()
    assert source.read_bytes() == before


def run(source: Path, output: Path, **kwargs: Any) -> dict[str, Any]:  # noqa: ANN401 - synthetic call overrides
    arguments = {
        "expected_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "source_locator": "https://example.invalid/gdp.xlsx",
        "source_vintage": gdp.VINTAGE,
        "observed_at": "2026-08-29T09:00:17Z",
        **kwargs,
    }
    return gdp.normalize_gdp(source, output, **arguments)


def test_complete_written_closure_and_determinism(tmp_path: Path) -> None:  # noqa: PLR0915 - independent complete closure assertions
    source = workbook(tmp_path / "source.xlsx")
    book = load_workbook(source)
    book["Contents"]["A34"] = "unselected context"
    book["Table 1"]["C12"] = 777
    book["Table 1"]["C27"] = -123
    book["Table 1"]["D27"] = 0
    book.save(source)
    book.close()
    before = source.read_bytes()
    first, second = tmp_path / "first", tmp_path / "second"
    receipt = run(source, first, dry_run=False)
    assert run(source, second, dry_run=False) == receipt
    assert {p.name for p in first.iterdir()} == {
        "MANIFEST.json",
        "gdp_facts.parquet",
        "field_lineage.parquet",
        "cell_dispositions.parquet",
    }
    for path in first.iterdir():
        assert path.read_bytes() == (second / path.name).read_bytes()
    assert source.read_bytes() == before
    assert json.loads((first / "MANIFEST.json").read_text()) == receipt
    for name, digest in receipt["output_sha256"].items():
        assert hashlib.sha256((first / name).read_bytes()).hexdigest() == digest
    facts = pq.read_table(first / "gdp_facts.parquet").to_pylist()
    lineage = pq.read_table(first / "field_lineage.parquet").to_pylist()
    dispositions = pq.read_table(first / "cell_dispositions.parquet").to_pylist()
    assert receipt["counts"] == {
        "facts": 60,
        "lineage": 900,
        "dispositions": len(dispositions),
    }
    assert len({row["record_id"] for row in facts}) == 60
    assert [row["amount"] for row in facts[:2]] == [Decimal(-123), Decimal(0)]
    assert facts[0]["period_end"] == date(2011, 6, 30)
    assert facts[-1]["period_end"] == date(2026, 3, 31)
    book = load_workbook(source)
    cells = {
        f"'{sheet.title}'!{cell.coordinate}": cell.value
        for sheet in book
        for row in sheet
        for cell in row
        if cell.value is not None
    }
    book.close()
    assert len(dispositions) == len(cells)
    assert {row["source_coordinate"] for row in dispositions} == set(cells)
    selected = {
        row["source_coordinate"]: row["record_id"]
        for row in dispositions
        if row["disposition"] == "selected"
    }
    assert selected == {row["source_coordinate"]: row["record_id"] for row in facts}
    for row in dispositions:
        assert str(json.loads(row["raw_value_json"])) == str(
            cells[row["source_coordinate"]]
        )
    for fact in facts:
        assert fact["currency"] is None
        assert fact["series_prefix"] == "SNEQ"
        assert fact["series_reference"] == "SG03AB01GE00S900"
        assert fact["adjustment"] == "actual_as_published"
        assert fact["price_basis"] == "current_prices"
        assert fact["unit"] == "$(million)"
        assert fact["rights_state"] == "not_evaluated"
        assert fact["scaling"] == "million"
        assert fact["amount"] == Decimal(str(cells[fact["source_coordinate"]]))
        entries = [row for row in lineage if row["record_id"] == fact["record_id"]]
        raw = json.loads(fact["raw_values_json"])
        assert len(entries) == len(raw) == 15
        assert {row["field"] for row in entries} == {
            key.split(":", 1)[0] for key in raw
        }
        assert {
            row["source_coordinate"] for row in entries if row["field"] == "period_end"
        } == {
            f"'Table 1'!{fact['source_coordinate'].split('!')[1][:-2]}6",
            "'Table 1'!A4",
        }
        for row in entries:
            assert row["lineage_id"] == fact["lineage_id"]
            assert row["source_object_sha256"] == fact["source_object_sha256"]
            assert row["source_locator"] == fact["source_locator"]
            expected = (
                "#,##0"
                if row["field"] == "source_number_format"
                else str(cells[row["source_coordinate"]])
            )
            assert row["raw_value"] == expected
            assert row["normalized_value"] == str(fact[row["field"]])
            key = (
                "period_end:range"
                if row["field"] == "period_end"
                and row["source_coordinate"] == "'Table 1'!A4"
                else row["field"]
            )
            assert raw[key] == [row["source_coordinate"], row["raw_value"]]


@pytest.mark.parametrize(
    ("sheet", "coordinate", "value"),
    [
        ("Table 1", "A1", "Table 2"),
        ("Table 1", "A3", "Seasonally adjusted current prices"),
        ("Table 1", "C6", "Jun-12"),
        ("Table 1", "D6", "Jun-11"),
        ("Table 1", "C27", None),
        ("Table 1", "C27", "-"),
        ("Table 1", "C27", True),
        ("Table 1", "C27", "=1+2"),
        ("Table 1", "C27", 1.5),
        ("Table 1", "BK27", 1),
        ("Table 1", "A32", "unexpected extent"),
        ("Table 2", "A3", "Actual current prices"),
        ("Contents", "A1", "Wrong release"),
        ("Contents", "A33", "19 June 2026"),
        *[("Table 1", key, "drift") for key in gdp.HEADERS],
    ],
)
def test_profile_drift_fails_before_output(
    tmp_path: Path,
    sheet: str,
    coordinate: str,
    value: str | float | None,
) -> None:
    source = workbook(tmp_path / "source.xlsx")
    book = load_workbook(source)
    book[sheet][coordinate] = value
    book.save(source)
    book.close()
    with pytest.raises(ValueError, match="gdp_source_contract"):
        run(source, tmp_path / "output")
    assert not (tmp_path / "output").exists()


@pytest.mark.parametrize("change", ["rename", "merge", "format"])
def test_structure_changes(tmp_path: Path, change: str) -> None:
    source = workbook(tmp_path / "source.xlsx")
    book = load_workbook(source)
    if change == "rename":
        book["Table 2"].title = "Other"
    elif change == "merge":
        book["Table 1"].unmerge_cells("C5:BJ5")
    else:
        book["Table 1"]["C27"].number_format = "0.0"
    book.save(source)
    book.close()
    with pytest.raises(ValueError, match="gdp_source_contract"):
        run(source, tmp_path / "output")


def test_path_hash_context_and_limits(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = workbook(tmp_path / "source.xlsx")
    output = tmp_path / "output"
    for kwargs, error in [
        ({"expected_sha256": "f" * 64}, "source_hash_mismatch"),
        ({"source_vintage": "other"}, "gdp_source_contract"),
    ]:
        with pytest.raises(ValueError, match=error):
            run(source, output, **kwargs)
    output.mkdir()
    with pytest.raises(ValueError, match="gdp_source_contract"):
        run(source, output)
    link = tmp_path / "link"
    try:
        link.symlink_to(source)
    except OSError:
        pytest.skip("symlink privilege unavailable")
    with pytest.raises(ValueError, match="gdp_source_contract"):
        run(link, tmp_path / "new")
    dangling = tmp_path / "dangling"
    dangling.symlink_to(tmp_path / "missing")
    with pytest.raises(ValueError, match="gdp_source_contract"):
        run(source, dangling)
    monkeypatch.setattr(gdp, "MAX_BYTES", source.stat().st_size - 1)
    with pytest.raises(ValueError, match="source_byte_limit"):
        run(source, tmp_path / "new")
    monkeypatch.setattr(gdp, "MAX_BYTES", source.stat().st_size)
    assert run(source, tmp_path / "new")["status"] == "planned"


def test_partial_write_is_not_complete(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = workbook(tmp_path / "source.xlsx")
    original = workbook_common.pq.write_table
    calls = 0

    def fail_second(*args: Any, **kwargs: Any) -> None:  # noqa: ANN401 - forwarding writer test double
        nonlocal calls
        calls += 1
        if calls == 2:
            message = "synthetic interruption"
            raise OSError(message)
        original(*args, **kwargs)

    monkeypatch.setattr(workbook_common.pq, "write_table", fail_second)
    with pytest.raises(OSError, match="synthetic interruption"):
        run(source, tmp_path / "output", dry_run=False)
    assert (tmp_path / "output" / "gdp_facts.parquet").exists()
    assert not (tmp_path / "output" / "MANIFEST.json").exists()


@given(st.integers(min_value=-(10**19), max_value=10**19))
def test_exact_integer_tokens(number: int) -> None:
    assert gdp._amount(str(number)) == number  # noqa: SLF001 - exact grammar property


@pytest.mark.parametrize(
    "token",
    ["", "-", "NA", "NaN", "Infinity", "1e3", "1.0", "+1", " 1", "1 ", "1" * 21],
)
def test_bad_numeric_tokens(token: str) -> None:
    with pytest.raises(ValueError, match="gdp_source_contract"):
        gdp._amount(token)  # noqa: SLF001 - lexical boundary


def test_period_headers_ignore_locale(monkeypatch: pytest.MonkeyPatch) -> None:
    with monkeypatch.context() as patch:
        patch.setattr(calendar, "month_abbr", ["localized"] * 13)
        importlib.reload(gdp)
        assert gdp.PERIODS[0] == ("Jun-11", date(2011, 6, 30))
        assert gdp.PERIODS[-1] == ("Mar-26", date(2026, 3, 31))
    importlib.reload(gdp)
