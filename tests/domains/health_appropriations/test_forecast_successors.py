"""Version-specific BEFU 2026 / HYEFU 2025 Health summary contracts."""

from __future__ import annotations

import hashlib
from pathlib import Path

import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook, load_workbook

from archive_govt_nz.domains.health_appropriations.forecast import (
    normalize_forecast_workbook,
)

PROFILES = [("befu-2026/v1", "BEFU-2026", 9), ("hyefu-2025/v1", "HYEFU-2025", 8)]


def _source(path: Path, health_row: int) -> str:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Core Crown Expense Tables"
    sheet.cell(health_row, 4, "Health")
    sheet.cell(health_row - 3, 4, "($millions)")
    for index, year in enumerate(range(2021, 2031)):
        sheet.cell(health_row - 4, 6 + index, str(year))
        sheet.cell(health_row - 3, 6 + index, "Actual" if index < 5 else "Forecast")
        sheet.cell(health_row, 6 + index, index - 1)
    sheet.cell(health_row + 2, 4, "Health expenses")
    sheet.cell(health_row + 2, 6, "=SUM(F8:F9)")
    workbook.create_sheet("Index")
    workbook.save(path)
    workbook.close()
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _run(source: Path, output: Path, profile: str, vintage: str) -> dict[str, object]:
    return normalize_forecast_workbook(
        source,
        output,
        expected_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
        profile=profile,
        observed_at="2026-08-29T09:00:17Z",
        source_vintage=vintage,
        source_locator="synthetic-expenses.xlsx",
    )


@pytest.mark.parametrize(("profile", "vintage", "health_row"), PROFILES)
def test_exact_successor_profile_retains_vintage_and_exclusions(
    tmp_path: Path, profile: str, vintage: str, health_row: int
) -> None:
    source = tmp_path / "source.xlsx"
    digest = _source(source, health_row)
    output = tmp_path / "out"
    receipt = _run(source, output, profile, vintage)
    assert receipt["status"] == "passed"
    assert receipt["profile"] == profile
    assert receipt["source_vintage"] == vintage
    facts = pq.read_table(output / "forecast_facts.parquet").to_pylist()
    assert [row["year"] for row in facts] == list(range(2021, 2031))
    assert [row["amount_type"] for row in facts] == ["Actual"] * 5 + ["Forecast"] * 5
    assert [row["amount"] for row in facts] == list(range(-1, 9))
    assert {row["source_vintage"] for row in facts} == {vintage}
    assert all(row["valid_time_start"] is None for row in facts)
    assert pq.read_table(output / "field_lineage.parquet").num_rows == 60
    cells = pq.read_table(output / "cell_dispositions.parquet").to_pylist()
    assert (
        next(row for row in cells if row["data_type"] == "f")["disposition"]
        == "preserved_only"
    )
    assert receipt["excluded_sheets"] == [
        {"sheet": "Index", "reason": "not_forecast_summary_sheet"}
    ]
    assert hashlib.sha256(source.read_bytes()).hexdigest() == digest


@pytest.mark.parametrize(("profile", "vintage", "health_row"), PROFILES)
@pytest.mark.parametrize(
    "change",
    ["vintage", "year", "numeric_year", "amount_type", "shift", "columns", "unit"],
)
def test_successor_drift_fails_before_output(
    tmp_path: Path, profile: str, vintage: str, health_row: int, change: str
) -> None:
    source = tmp_path / "source.xlsx"
    _source(source, health_row)
    workbook = load_workbook(source)
    sheet = workbook["Core Crown Expense Tables"]
    if change == "vintage":
        vintage = "old-vintage"
    elif change in ("year", "numeric_year"):
        for index, year in enumerate(range(2021, 2031)):
            sheet.cell(
                health_row - 4, 6 + index, str(year - 1) if change == "year" else year
            )
    elif change == "amount_type":
        sheet.cell(health_row - 3, 10, "Forecast")
    elif change in ("columns", "unit"):
        selected_rows = (
            (health_row - 4, health_row - 3, health_row)
            if change == "columns"
            else (health_row - 4, health_row - 3)
        )
        cells = [
            (cell.row, cell.column, cell.value)
            for row in selected_rows
            for cell in sheet[row]
            if cell.value is not None and (change == "unit" or cell.column >= 6)
        ]
        for row, column, _value in cells:
            sheet.cell(row, column).value = None
        for row, column, value in cells:
            sheet.cell(row - (change == "unit"), column + (change == "columns"), value)
    else:
        sheet.insert_rows(1)
    workbook.save(source)
    workbook.close()
    with pytest.raises(ValueError, match="forecast_vintage_contract"):
        _run(source, tmp_path / "out", profile, vintage)
    assert not (tmp_path / "out").exists()


@pytest.mark.parametrize(("profile", "vintage", "health_row"), PROFILES)
def test_successor_formula_stays_rejected_not_evaluated(
    tmp_path: Path, profile: str, vintage: str, health_row: int
) -> None:
    source = tmp_path / "source.xlsx"
    _source(source, health_row)
    workbook = load_workbook(source)
    workbook["Core Crown Expense Tables"].cell(health_row, 6, "=1+1")
    workbook.save(source)
    workbook.close()
    receipt = _run(source, tmp_path / "out", profile, vintage)
    assert receipt["status"] == "partial"
    assert pq.read_table(tmp_path / "out/forecast_facts.parquet").num_rows == 9
    cells = pq.read_table(tmp_path / "out/cell_dispositions.parquet").to_pylist()
    selected = next(
        row
        for row in cells
        if row["source_coordinate"] == f"'Core Crown Expense Tables'!F{health_row}"
    )
    assert selected["disposition"] == "rejected"
    assert selected["reason"] == "formula_not_evaluated"
