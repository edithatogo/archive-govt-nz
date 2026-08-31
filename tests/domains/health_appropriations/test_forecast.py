"""Literal forecast summaries retain vintage, source cells and exclusions."""

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import tempfile
from decimal import Decimal
from pathlib import Path
from typing import Any

import pyarrow.parquet as pq
import pytest
from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import Workbook, load_workbook

from archive_govt_nz.domains.health_appropriations import forecast, workbook_common
from archive_govt_nz.domains.health_appropriations.forecast import (
    normalize_forecast_workbook,
)

SHEETS = {"befu": "Core Crown Expense Tables", "hyefu": "Expense Tables"}


def _source(path: Path, profile: str = "befu", *, shift: int = 0) -> str:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEETS[profile]
    for coordinate, value in {
        "D2": 2024,
        "E2": 2025,
        "B3": "($millions)",
        "D3": "Actual",
        "E3": "Forecast",
        "B5": "Health",
        "D5": 0,
        "E5": -5,
        "B6": "Education",
        "D6": 200,
        "B9": "Health expenses",
        "D9": "=SUM(D5:D5)",
    }.items():
        sheet[coordinate] = value
    if shift:
        sheet.insert_rows(1, shift)
        sheet.insert_cols(1, shift)
    workbook.create_sheet("Index")
    workbook.save(path)
    workbook.close()
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _run(
    source: Path, output: Path, digest: str, profile: str = "befu"
) -> dict[str, Any]:
    return normalize_forecast_workbook(
        source,
        output,
        expected_sha256=digest,
        profile=profile,
        observed_at="2026-08-30T07:15:00Z",
        source_vintage=f"{profile}-2025",
        source_locator=f"data/raw/{profile}-expense.xlsx",
    )


@pytest.mark.parametrize("profile", ["befu", "hyefu"])
@pytest.mark.parametrize("shift", [0, 3])
def test_literal_summaries_have_semantic_layout_and_lineage(
    tmp_path: Path, profile: str, shift: int
) -> None:
    source = tmp_path / "object"
    digest = _source(source, profile, shift=shift)
    receipt = _run(source, tmp_path / "one", digest, profile)
    assert receipt["status"] == "passed"
    assert receipt["counts"]["normalized"] == 2
    assert receipt["counts"]["rejected"] == 0
    assert receipt["counts"]["preserved_only"] == 4
    assert receipt["counts"]["context"] == 6
    facts = pq.read_table(tmp_path / "one/forecast_facts.parquet").to_pylist()
    assert [row["year"] for row in facts] == [2024, 2025]
    assert [row["amount"] for row in facts] == [Decimal(0), Decimal(-5)]
    assert [row["amount_type"] for row in facts] == ["Actual", "Forecast"]
    assert {row["unit"] for row in facts} == {"NZD_millions"}
    assert {row["source_object_sha256"] for row in facts} == {digest}
    assert {row["source_vintage"] for row in facts} == {f"{profile}-2025"}
    assert all(row["valid_time_start"] is None for row in facts)
    assert all(row["donor_table"] is None for row in facts)
    assert all(row["rights_state"] == "not_evaluated" for row in facts)
    assert all(
        row["quality_flags"] == ["financial_year_basis_unverified"] for row in facts
    )
    lineage = pq.read_table(tmp_path / "one/field_lineage.parquet").to_pylist()
    assert len(lineage) == 12
    assert {row["field"] for row in lineage} == {
        "year",
        "amount",
        "amount_type",
        "functional_classification",
        "measure",
        "unit",
    }
    amounts = [row for row in lineage if row["field"] == "amount"]
    coordinate = "G8" if shift else "D5"
    assert amounts[0]["source_coordinate"] == f"'{SHEETS[profile]}'!{coordinate}"
    assert amounts[0]["raw_value"] == "0"
    assert amounts[0]["normalized_value"] == "0.000"
    cells = pq.read_table(tmp_path / "one/cell_dispositions.parquet").to_pylist()
    formula = next(row for row in cells if row["data_type"] == "f")
    assert formula["disposition"] == "preserved_only"
    assert formula["reason"] == "outside_literal_health_summary"
    assert receipt["excluded_sheets"] == [
        {"sheet": "Index", "reason": "not_forecast_summary_sheet"}
    ]
    assert json.loads((tmp_path / "one/MANIFEST.json").read_text()) == receipt
    assert _run(source, tmp_path / "two", digest, profile) == receipt
    for path in (tmp_path / "one").iterdir():
        assert path.read_bytes() == (tmp_path / "two" / path.name).read_bytes()
    assert hashlib.sha256(source.read_bytes()).hexdigest() == digest


@pytest.mark.parametrize(
    ("coordinate", "value", "reason"),
    [
        ("D5", None, "invalid_amount"),
        ("D5", True, "invalid_amount"),
        ("D5", "bad", "invalid_amount"),
        ("D5", "1.0001", "invalid_amount"),
        ("D5", "=1+1", "formula_not_evaluated"),
        ("D5", "#DIV/0!", "spreadsheet_error"),
    ],
)
def test_rejected_amounts_remain_visible(
    tmp_path: Path, coordinate: str, value: str | float | None, reason: str
) -> None:
    source = tmp_path / "book.xlsx"
    _source(source)
    workbook = load_workbook(source)
    workbook[SHEETS["befu"]][coordinate] = value
    workbook.save(source)
    workbook.close()
    receipt = _run(
        source, tmp_path / "out", hashlib.sha256(source.read_bytes()).hexdigest()
    )
    assert receipt["status"] == "partial"
    assert receipt["counts"]["normalized"] == 1
    assert receipt["counts"]["rejected"] == 1
    cells = pq.read_table(tmp_path / "out/cell_dispositions.parquet").to_pylist()
    rejected = next(row for row in cells if row["disposition"] == "rejected")
    assert rejected["reason"] == reason
    assert rejected["source_coordinate"] == f"'{SHEETS['befu']}'!D5"
    assert rejected["record_id"] is None


@pytest.mark.parametrize(
    ("coordinate", "value", "reason"),
    [
        ("B5", "Health expenses", "health_summary_label"),
        ("B6", "Health", "health_summary_label"),
        ("B3", "$thousands", "summary_unit"),
        ("D2", "year", "summary_years"),
        ("D2", True, "summary_years"),
        ("D2", "=2024", "summary_years"),
        ("D2", 2025, "summary_years"),
        ("D2", 2023, "summary_years"),
        ("E2", None, "summary_years"),
        ("E2", 2025.5, "summary_years"),
        ("D3", "Estimated Actual", "summary_amount_types"),
        ("D3", "=1", "summary_amount_types"),
        ("D3", "Forecast", "summary_amount_types"),
        ("F5", 99, "unlabelled_summary_value"),
    ],
)
def test_layout_drift_fails_before_output(
    tmp_path: Path, coordinate: str, value: str | float | None, reason: str
) -> None:
    source = tmp_path / "book.xlsx"
    _source(source)
    workbook = load_workbook(source)
    sheet = workbook[SHEETS["befu"]]
    sheet[coordinate] = value
    if coordinate == "D3" and value == "Forecast":
        sheet["E3"] = "Actual"
    workbook.save(source)
    workbook.close()
    with pytest.raises(ValueError, match=reason):
        _run(source, tmp_path / "out", hashlib.sha256(source.read_bytes()).hexdigest())
    assert not (tmp_path / "out").exists()


def test_profile_hash_and_output_collision(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source)
    with pytest.raises(ValueError, match="unsupported_forecast_profile"):
        _run(source, tmp_path / "out", digest, "unknown")
    with pytest.raises(ValueError, match="missing_forecast_sheet"):
        _run(source, tmp_path / "out", digest, "hyefu")
    with pytest.raises(ValueError, match="source_hash_mismatch"):
        _run(source, tmp_path / "out", "a" * 64)
    (tmp_path / "out").mkdir()
    with pytest.raises(FileExistsError):
        _run(source, tmp_path / "out", digest)
    assert list((tmp_path / "out").iterdir()) == []


@pytest.mark.parametrize("mode", ["unit_first_row", "noncontiguous_columns"])
def test_context_boundaries(tmp_path: Path, mode: str) -> None:
    source = tmp_path / "book.xlsx"
    _source(source)
    workbook = load_workbook(source)
    sheet = workbook[SHEETS["befu"]]
    if mode == "unit_first_row":
        sheet["B3"] = None
        sheet["B1"] = "($millions)"
    else:
        sheet["E2"] = None
        sheet["F2"] = 2025
    workbook.save(source)
    workbook.close()
    with pytest.raises(ValueError, match="summary_years"):
        _run(source, tmp_path / "out", hashlib.sha256(source.read_bytes()).hexdigest())
    assert not (tmp_path / "out").exists()


def test_all_forecast_and_all_rejected_are_explicit(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    _source(source)
    workbook = load_workbook(source)
    sheet = workbook[SHEETS["befu"]]
    sheet["D3"] = "Forecast"
    workbook.save(source)
    receipt = _run(
        source, tmp_path / "good", hashlib.sha256(source.read_bytes()).hexdigest()
    )
    assert receipt["status"] == "passed"
    sheet["D5"] = None
    sheet["E5"] = None
    workbook.save(source)
    workbook.close()
    receipt = _run(
        source, tmp_path / "bad", hashlib.sha256(source.read_bytes()).hexdigest()
    )
    assert receipt["status"] == "partial"
    assert receipt["counts"]["normalized"] == 0
    assert receipt["counts"]["rejected"] == 2
    assert pq.read_table(tmp_path / "bad/forecast_facts.parquet").num_rows == 0


def test_cap_and_failed_write_preserve_original(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source)
    monkeypatch.setattr(forecast, "_MAX_SOURCE_BYTES", source.stat().st_size - 1)
    with pytest.raises(ValueError, match="source_byte_limit"):
        _run(source, tmp_path / "out", digest)
    monkeypatch.setattr(forecast, "_MAX_SOURCE_BYTES", source.stat().st_size)

    def fail_write(*_args: object, **_kwargs: object) -> None:
        message = "injected write failure"
        raise OSError(message)

    monkeypatch.setattr(workbook_common.pq, "write_table", fail_write)
    with pytest.raises(OSError, match="injected write failure"):
        _run(source, tmp_path / "out", digest)
    assert not (tmp_path / "out/MANIFEST.json").exists()
    assert hashlib.sha256(source.read_bytes()).hexdigest() == digest


@pytest.mark.parametrize(
    ("bad_amount", "status", "returncode"), [(False, "passed", 0), (True, "partial", 2)]
)
def test_command_reports_extraction_state(
    tmp_path: Path, *, bad_amount: bool, status: str, returncode: int
) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source)
    if bad_amount:
        workbook = load_workbook(source)
        workbook[SHEETS["befu"]]["D5"] = "bad"
        workbook.save(source)
        workbook.close()
        digest = hashlib.sha256(source.read_bytes()).hexdigest()
    result = subprocess.run(
        [
            sys.executable,
            "tools/build_health_forecast.py",
            "--source",
            str(source),
            "--expected-sha256",
            digest,
            "--profile",
            "befu",
            "--output-dir",
            str(tmp_path / "out"),
            "--observed-at",
            "2026-08-30T07:15:00Z",
            "--source-vintage",
            "BEFU-2025",
            "--source-locator",
            "data/raw/befu25-data-expense-tables.xlsx",
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == returncode, result.stderr
    assert json.loads(result.stdout)["status"] == status


@settings(
    max_examples=12,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture],
)
@given(
    start_year=st.integers(min_value=1990, max_value=2040),
    width=st.integers(min_value=2, max_value=6),
    amount=st.integers(min_value=-1_000_000, max_value=1_000_000),
)
def test_generated_series_preserve_every_value(
    tmp_path: Path, start_year: int, width: int, amount: int
) -> None:
    with tempfile.TemporaryDirectory(dir=tmp_path) as temporary:
        root = Path(temporary)
        source = root / "book.xlsx"
        _source(source)
        workbook = load_workbook(source)
        sheet = workbook[SHEETS["befu"]]
        for index in range(width):
            sheet.cell(2, 4 + index, start_year + index)
            sheet.cell(3, 4 + index, "Actual" if index == 0 else "Forecast")
            sheet.cell(5, 4 + index, amount + index)
        workbook.save(source)
        workbook.close()
        digest = hashlib.sha256(source.read_bytes()).hexdigest()
        receipt = _run(source, root / "out", digest)
        facts = pq.read_table(root / "out/forecast_facts.parquet").to_pylist()
        assert receipt["status"] == "passed"
        assert receipt["counts"]["normalized"] == width
        assert [row["year"] for row in facts] == list(
            range(start_year, start_year + width)
        )
        assert [row["amount"] for row in facts] == [
            Decimal(amount + index) for index in range(width)
        ]
        assert len({row["record_id"] for row in facts}) == width
        assert pq.read_table(root / "out/field_lineage.parquet").num_rows == width * 6
        assert hashlib.sha256(source.read_bytes()).hexdigest() == digest
