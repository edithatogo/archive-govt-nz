"""Explicit no-write forecast inspection preserves legacy write behavior."""

from __future__ import annotations

import hashlib
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pytest
from openpyxl import load_workbook
from tests.domains.health_appropriations.test_forecast import _source as legacy_source
from tests.domains.health_appropriations.test_forecast_successors import (
    _source as successor_source,
)

from archive_govt_nz.domains.health_appropriations import forecast


@pytest.fixture(
    params=[
        ("befu", "befu-2025", None),
        ("hyefu", "hyefu-2025", None),
        ("befu-2026/v1", "BEFU-2026", 9),
        ("hyefu-2025/v1", "HYEFU-2025", 8),
    ]
)
def case(tmp_path: Path, request: pytest.FixtureRequest) -> tuple[Path, dict[str, Any]]:
    profile, vintage, row = request.param
    source = tmp_path / "source.xlsx"
    if row is None:
        digest = legacy_source(source, profile)
    else:
        digest = successor_source(source, row)
    return source, {
        "expected_sha256": digest,
        "profile": profile,
        "source_vintage": vintage,
        "source_locator": "synthetic-expenses.xlsx",
        "observed_at": "2026-08-29T09:00:17Z",
    }


def test_explicit_preflight_never_calls_writer(
    tmp_path: Path,
    case: tuple[Path, dict[str, Any]],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source, context = case
    before = source.read_bytes()
    output = tmp_path / "missing-parent" / "output"

    def forbidden(*_args: object, **_kwargs: object) -> None:
        pytest.fail("preflight called writer")

    monkeypatch.setattr(forecast, "write_workbook_outputs", forbidden)
    monkeypatch.setattr(
        forecast, "pa", SimpleNamespace(Table=SimpleNamespace(from_pylist=forbidden))
    )
    receipt = forecast.normalize_forecast_workbook(
        source, output, **context, dry_run=True
    )
    assert receipt["status"] == "planned"
    assert receipt["preflight_scope"] == "source_validation_only"
    assert receipt["rights_state"] == "not_evaluated"
    assert "output_sha256" not in receipt
    assert not output.parent.exists()
    assert source.read_bytes() == before


@pytest.mark.parametrize("target", ["file", "directory", "source"])
def test_preflight_does_not_validate_or_modify_output_location(
    tmp_path: Path, case: tuple[Path, dict[str, Any]], target: str
) -> None:
    source, context = case
    output = tmp_path / "existing"
    if target == "file":
        output.write_bytes(b"retain")
        retained = output
    elif target == "directory":
        output.mkdir()
        retained = output / "retained"
        retained.write_bytes(b"retain")
    else:
        output = source
        retained = source
    before = retained.read_bytes()
    plan = forecast.normalize_forecast_workbook(source, output, **context, dry_run=True)
    assert plan["preflight_scope"] == "source_validation_only"
    assert plan["status"] == "planned"
    assert retained.read_bytes() == before
    if target == "directory":
        assert {p.name for p in output.iterdir()} == {"retained"}


@pytest.mark.parametrize("change", ["hash", "layout"])
def test_invalid_source_cannot_create_output_during_preflight(
    tmp_path: Path, case: tuple[Path, dict[str, Any]], change: str
) -> None:
    source, context = case
    expected = "source_hash_mismatch"
    if change == "hash":
        context["expected_sha256"] = "a" * 64
    else:
        workbook = load_workbook(source)
        for row in workbook.worksheets[0].iter_rows():
            for cell in row:
                if cell.value == "Health":
                    cell.value = "Other"
        workbook.save(source)
        workbook.close()
        context["expected_sha256"] = hashlib.sha256(source.read_bytes()).hexdigest()
        expected = "health_summary_label"
    before = source.read_bytes()
    output = tmp_path / "missing" / "output"
    with pytest.raises(ValueError, match=expected):
        forecast.normalize_forecast_workbook(source, output, **context, dry_run=True)
    assert source.read_bytes() == before
    assert not output.parent.exists()


def test_preflight_propagates_interrupts_without_creating_state(
    tmp_path: Path,
    case: tuple[Path, dict[str, Any]],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source, context = case

    def interrupt(*_args: object, **_kwargs: object) -> None:
        raise KeyboardInterrupt

    monkeypatch.setattr(forecast, "verified_snapshot", interrupt)
    output = tmp_path / "absent"
    with pytest.raises(KeyboardInterrupt):
        forecast.normalize_forecast_workbook(source, output, **context, dry_run=True)
    assert not output.exists()


def test_legacy_default_and_explicit_write_are_identical(
    tmp_path: Path, case: tuple[Path, dict[str, Any]]
) -> None:
    source, context = case
    first, second = tmp_path / "first", tmp_path / "second"
    receipt = forecast.normalize_forecast_workbook(source, first, **context)
    assert (
        forecast.normalize_forecast_workbook(source, second, **context, dry_run=False)
        == receipt
    )
    assert {p.name: p.read_bytes() for p in first.iterdir()} == {
        p.name: p.read_bytes() for p in second.iterdir()
    }
    plan = forecast.normalize_forecast_workbook(
        source, tmp_path / "plan", **context, dry_run=True
    )
    assert plan.pop("preflight_scope") == "source_validation_only"
    assert {
        **plan,
        "status": "passed",
        "output_sha256": receipt["output_sha256"],
    } == receipt


@pytest.mark.parametrize("flag", [None, 0, 1, "false", [], {}])
def test_nonboolean_flag_rejected_before_source_access(
    tmp_path: Path, flag: object
) -> None:
    with pytest.raises(ValueError, match=r"^forecast_dry_run_type$"):
        forecast.normalize_forecast_workbook(
            tmp_path / "absent",
            tmp_path / "output",
            expected_sha256="a" * 64,
            profile="befu",
            source_vintage="BEFU-2025",
            source_locator="synthetic.xlsx",
            observed_at="2026-08-29T00:00:00Z",
            dry_run=flag,  # type: ignore[arg-type]
        )


def test_rejected_amount_is_not_a_successful_plan(
    tmp_path: Path, case: tuple[Path, dict[str, Any]]
) -> None:
    source, context = case
    workbook = load_workbook(source)
    sheet = workbook.worksheets[0]
    coordinates = {
        "befu": "D5",
        "hyefu": "D5",
        "befu-2026/v1": "F9",
        "hyefu-2025/v1": "F8",
    }
    sheet[coordinates[context["profile"]]] = "=1+1"
    workbook.save(source)
    workbook.close()
    context["expected_sha256"] = hashlib.sha256(source.read_bytes()).hexdigest()
    before = source.read_bytes()
    output = tmp_path / "no-state"
    receipt = forecast.normalize_forecast_workbook(
        source, output, **context, dry_run=True
    )
    assert receipt["status"] == "partial"
    assert receipt["counts"]["rejected"] == 1  # type: ignore[index]
    assert not output.exists()
    assert source.read_bytes() == before
