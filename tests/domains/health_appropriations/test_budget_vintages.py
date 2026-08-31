"""Synthetic characterization of the full Budget expenditure layout by vintage.

Headers and year/type windows model observed workbook contracts; every amount,
identifier, scope and appropriation below is invented, not copied source data.
These tests characterize existing behavior, not a new extraction implementation.
"""

from __future__ import annotations

import hashlib
import json
from decimal import Decimal
from pathlib import Path
from typing import Any

import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from archive_govt_nz.domains.health_appropriations.appropriation_analysis import (
    analyze_appropriations,
)
from archive_govt_nz.domains.health_appropriations.budget import (
    normalize_budget_workbook,
)

HEADERS = (
    "Department",
    "Vote",
    "App ID",
    "Parent ID",
    "Appropriation Name",
    "Category Name",
    "Group Type",
    "Appropriation or Category Type",
    "Restriction Type",
    "Functional Classification",
    "Amount $000",
    "Year",
    "Amount Type",
    "Periodicity",
    "Current Scope",
    "M Number",
    "Portfolio Name",
)
NORMALIZED_FIELDS = {
    "Department": "department",
    "Appropriation Name": "appropriation_name",
    "Functional Classification": "functional_classification",
    "Amount $000": "amount",
    "Year": "year",
    "Amount Type": "amount_type",
    "Portfolio Name": "portfolio_name",
}


def _rows(vintage: int) -> list[list[object]]:
    periods = [
        *((year, "Actuals") for year in range(vintage - 4, vintage)),
        (vintage, "Estimated Actual"),
        (vintage + 1, "Main Estimates"),
        (2023, "Actuals"),
    ]
    amounts = ["0.000", "1.125", "-2.750", "10.010", "100.125", "200.250", "-0.125"]
    return [
        [
            "Synthetic Department",
            "Health",
            f"synthetic-{vintage}-{index}",
            "synthetic-parent",
            "Synthetic care",
            "Synthetic category",
            "Synthetic group",
            "Synthetic appropriation type",
            "Synthetic restriction",
            "Health",
            amount,
            year,
            amount_type,
            "Annual",
            "Synthetic scope only",
            "synthetic-minister",
            "Synthetic portfolio",
        ]
        for index, ((year, amount_type), amount) in enumerate(
            zip(periods, amounts, strict=True)
        )
    ]


def _build(tmp_path: Path, vintage: int) -> tuple[Path, str, list[list[object]]]:
    source = tmp_path / f"synthetic-budget-{vintage}.xlsx"
    rows = _rows(vintage)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Raw Data"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(source)
    workbook.close()
    return source, hashlib.sha256(source.read_bytes()).hexdigest(), rows


def _normalize(
    source: Path, digest: str, vintage: int, output: Path
) -> dict[str, object]:
    return normalize_budget_workbook(
        source,
        output,
        expected_sha256=digest,
        observed_at="2026-08-31T00:00:00Z",
        source_vintage=f"Budget-{vintage}",
        source_locator=f"synthetic/budget-{vintage}.xlsx",
    )


@pytest.mark.parametrize(
    ("vintage", "expected_periods"),
    [
        (
            2025,
            [
                (2021, "Actuals"),
                (2022, "Actuals"),
                (2023, "Actuals"),
                (2024, "Actuals"),
                (2025, "Estimated Actual"),
                (2026, "Main Estimates"),
                (2023, "Actuals"),
            ],
        ),
        (
            2026,
            [
                (2022, "Actuals"),
                (2023, "Actuals"),
                (2024, "Actuals"),
                (2025, "Actuals"),
                (2026, "Estimated Actual"),
                (2027, "Main Estimates"),
                (2023, "Actuals"),
            ],
        ),
    ],
)
def test_full_layout_preserves_values_all_cell_lineage_and_rebuild(
    tmp_path: Path, vintage: int, expected_periods: list[tuple[int, str]]
) -> None:
    source, digest, rows = _build(tmp_path, vintage)
    original = source.read_bytes()
    first = tmp_path / "first"
    receipt = _normalize(source, digest, vintage, first)
    assert receipt["status"] == "passed"
    assert receipt["counts"] == {
        "input": 7,
        "normalized": 7,
        "out_of_scope": 0,
        "blank": 0,
        "rejected": 0,
    }
    facts = pq.read_table(first / "budget_facts.parquet").to_pylist()
    assert [(fact["year"], fact["amount_type"]) for fact in facts] == expected_periods
    lineage = pq.read_table(first / "field_lineage.parquet").to_pylist()
    assert len(lineage) == 7 * 17
    assert len({fact["record_id"] for fact in facts}) == 7
    for source_row, (fact, raw) in enumerate(zip(facts, rows, strict=True), 2):
        values = dict(zip(HEADERS, raw, strict=True))
        assert json.loads(fact["raw_values_json"]) == values
        assert fact["amount"] == Decimal(str(values["Amount $000"]))
        assert fact["year"] == values["Year"]
        assert fact["amount_type"] == values["Amount Type"]
        assert fact["source_vintage"] == f"Budget-{vintage}"
        assert fact["source_object_sha256"] == digest
        assert fact["valid_time_start"] is None
        assert fact["quality_flags"] == ["financial_year_basis_unverified"]
        cells = [item for item in lineage if item["record_id"] == fact["record_id"]]
        assert len(cells) == 17
        for column, (cell, header) in enumerate(zip(cells, HEADERS, strict=True), 1):
            field = NORMALIZED_FIELDS.get(header, f"raw:{header}")
            assert cell["field"] == field
            assert cell["source_coordinate"] == (
                f"'Raw Data'!{get_column_letter(column)}{source_row}"
            )
            assert cell["raw_value"] == str(values[header])
            assert cell["normalized_value"] == str(fact.get(field, values[header]))
            assert cell["source_object_sha256"] == digest
            assert cell["source_locator"] == f"synthetic/budget-{vintage}.xlsx"
            assert cell["lineage_id"] == fact["lineage_id"]
    second = tmp_path / "second"
    assert _normalize(source, digest, vintage, second) == receipt
    assert {path.name for path in first.iterdir()} == {
        path.name for path in second.iterdir()
    }
    for path in first.iterdir():
        assert path.read_bytes() == (second / path.name).read_bytes()
    assert source.read_bytes() == original


def test_combined_vintages_keep_overlapping_years_types_and_inputs_separate(
    tmp_path: Path,
) -> None:
    facts: list[dict[str, Any]] = []
    digests = {}
    for vintage in (2025, 2026):
        source, digest, _ = _build(tmp_path, vintage)
        output = tmp_path / str(vintage)
        _normalize(source, digest, vintage, output)
        digests[f"Budget-{vintage}"] = digest
        facts.extend(pq.read_table(output / "budget_facts.parquet").to_pylist())
    result = analyze_appropriations(facts, breakdown_year=2026)
    assert len(result["trends"]) == 12
    assert analyze_appropriations(list(reversed(facts)), breakdown_year=2026) == result
    for group in result["trends"]:
        selected = [
            fact
            for fact in facts
            if all(
                fact[key] == group[key]
                for key in (
                    "source_object_sha256",
                    "source_vintage",
                    "year",
                    "amount_type",
                    "functional_classification",
                )
            )
        ]
        assert selected
        assert group["source_object_sha256"] == digests[group["source_vintage"]]
        assert group["input_record_ids"] == sorted(row["record_id"] for row in selected)
        assert Decimal(group["total_amount_thousands"]) == sum(
            (row["amount"] for row in selected), Decimal(0)
        )
    shared = [group for group in result["trends"] if group["year"] == 2023]
    assert {
        (group["source_vintage"], group["total_amount_thousands"]) for group in shared
    } == {
        ("Budget-2025", "-2.875"),
        ("Budget-2026", "1.000"),
    }
    assert {
        (group["source_vintage"], group["amount_type"])
        for group in result["trends"]
        if group["year"] == 2025
    } == {
        ("Budget-2025", "Estimated Actual"),
        ("Budget-2026", "Actuals"),
    }
    assert len(result["breakdown"]) == 1
    assert result["breakdown"][0]["source_vintage"] == "Budget-2026"
    assert result["breakdown"][0]["total_amount_thousands"] == "100.125"
    assert {
        record_id
        for group in result["trends"]
        for record_id in group["input_record_ids"]
    } == {fact["record_id"] for fact in facts}
