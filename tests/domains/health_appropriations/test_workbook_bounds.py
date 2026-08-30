"""Fail-closed workbook inventory bounds preserve the original package."""

from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

import pytest
from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import Workbook

from archive_govt_nz.domains.health_appropriations import formats


@pytest.mark.parametrize(
    "name",
    [
        "/absolute",
        "../escape",
        r"..\escape",
        "C:relative",
        "C:/absolute",
        r"\\server\share",
        "xl//sheet.xml",
        "xl/./sheet.xml",
        "xl/../../escape",
    ],
)
def test_inventory_rejects_unsafe_paths_on_every_platform(
    tmp_path: Path, name: str
) -> None:
    path = tmp_path / "unsafe.xlsx"
    with ZipFile(path, "w") as package:
        package.writestr(name, b"x")
    before = path.read_bytes()
    with pytest.raises(ValueError, match=r"^unsafe_workbook_member$"):
        formats.inventory_workbook(path)
    assert path.read_bytes() == before


@settings(max_examples=20, deadline=None)
@given(row=st.integers(1, 20), column=st.integers(1, 20), delta=st.integers(-1, 1))
def test_scan_budget_matches_rectangular_extent(
    row: int, column: int, delta: int
) -> None:
    with TemporaryDirectory() as directory, pytest.MonkeyPatch.context() as patch:
        path = Path(directory) / "bounded.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.cell(row=row, column=column, value="=1+2")
        workbook.save(path)
        workbook.close()
        before = path.read_bytes()
        patch.setattr(formats, "_MAX_SCAN_CELLS", row * column + delta)
        if delta < 0:
            with pytest.raises(ValueError, match=r"^workbook_cell_scan_limit$"):
                formats.inventory_workbook(path)
        else:
            assert formats.inventory_workbook(path)["kind"] == "xlsx"
        assert path.read_bytes() == before


def test_inventory_rejects_duplicate_parts(tmp_path: Path) -> None:
    path = tmp_path / "duplicate.xlsx"
    with ZipFile(path, "w") as package:
        package.writestr("xl/workbook.xml", b"first")
        with pytest.warns(UserWarning, match="Duplicate name"):
            package.writestr("xl/workbook.xml", b"second")
    before = path.read_bytes()
    with pytest.raises(ValueError, match=r"^duplicate_workbook_member$"):
        formats.inventory_workbook(path)
    assert path.read_bytes() == before


def test_inventory_bounds_cumulative_sparse_cell_scan(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "sparse.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["B2"] = "=1+2"
    workbook.create_sheet("Second")["B2"] = 7
    workbook.save(path)
    workbook.close()
    before = path.read_bytes()
    monkeypatch.setattr(formats, "_MAX_SCAN_CELLS", 7)
    with pytest.raises(ValueError, match=r"^workbook_cell_scan_limit$"):
        formats.inventory_workbook(path)
    monkeypatch.setattr(formats, "_MAX_SCAN_CELLS", 8)
    result = formats.inventory_workbook(path)
    sheets = result["sheets"]
    assert isinstance(sheets, list)
    assert len(sheets) == 2
    assert path.read_bytes() == before
