"""Raw Budget workbook extraction, provenance and failure contracts."""

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
from decimal import Decimal
from pathlib import Path

import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook

from archive_govt_nz.domains.health_appropriations import budget
from archive_govt_nz.domains.health_appropriations.budget import (
    normalize_budget_workbook,
)

HEADERS = [
    "Vote",
    "Year",
    "Department",
    "Appropriation Name",
    "Functional Classification",
    "Amount $000",
    "Amount Type",
    "Portfolio Name",
    "Current Scope",
]
ROW = [
    "Health",
    2025,
    "Health",
    "Care",
    "Health",
    123,
    "Main Estimates",
    "Health",
    "scope",
]


def _source(
    path: Path, rows: list[list[object]], headers: list[object] | None = None
) -> str:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Raw Data"
    sheet.append(HEADERS if headers is None else headers)
    for row in rows:
        sheet.append(row)
    workbook.create_sheet("Explanation")
    workbook.save(path)
    workbook.close()
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _run(source: Path, output: Path, digest: str) -> dict[str, object]:
    return normalize_budget_workbook(
        source,
        output,
        expected_sha256=digest,
        observed_at="2026-08-30T00:00:00Z",
        source_vintage="Budget-2025",
        source_locator="data/raw/b25-expenditure-data.xlsx",
    )


def test_raw_budget_retains_rows_lineage_and_original_bytes(tmp_path: Path) -> None:
    source = tmp_path / "original"
    digest = _source(source, [ROW, ["Education", *ROW[1:]], [], ROW])
    receipt = _run(source, tmp_path / "one", digest)
    assert receipt["status"] == "passed"
    assert receipt["counts"] == {
        "input": 4,
        "normalized": 2,
        "out_of_scope": 1,
        "blank": 1,
        "rejected": 0,
    }
    facts = pq.read_table(tmp_path / "one/budget_facts.parquet").to_pylist()
    assert len({row["record_id"] for row in facts}) == 2
    assert facts[0]["amount"] == Decimal("123.000")
    assert facts[0]["year"] == 2025
    assert facts[0]["valid_time_start"] is None
    assert facts[0]["donor_table"] is None
    assert facts[0]["donor_row_number"] is None
    assert facts[0]["rights_state"] == "not_evaluated"
    assert facts[0]["quality_flags"] == ["financial_year_basis_unverified"]
    assert json.loads(facts[0]["raw_values_json"])["Current Scope"] == "scope"
    lineage = pq.read_table(tmp_path / "one/field_lineage.parquet").to_pylist()
    assert len(lineage) == 18
    assert {row["source_object_sha256"] for row in lineage} == {digest}
    amount = next(row for row in lineage if row["field"] == "amount")
    assert amount["source_coordinate"] == "'Raw Data'!F2"
    assert amount["raw_value"] == "123"
    assert amount["normalized_value"] == "123.000"
    assert receipt["excluded_sheets"] == [
        {"sheet": "Explanation", "reason": "not_budget_raw_data"}
    ]
    assert json.loads((tmp_path / "one/MANIFEST.json").read_text()) == receipt
    assert _run(source, tmp_path / "two", digest) == receipt
    for path in (tmp_path / "one").iterdir():
        assert path.read_bytes() == (tmp_path / "two" / path.name).read_bytes()
    assert hashlib.sha256(source.read_bytes()).hexdigest() == digest


@pytest.mark.parametrize(
    ("column", "value", "reason"),
    [
        (0, None, "invalid_vote"),
        (0, "=1", "invalid_vote"),
        (1, 2025.5, "invalid_year"),
        (1, True, "invalid_year"),
        (5, True, "invalid_amount"),
        (5, "wrong", "invalid_amount"),
        (5, "1.0001", "invalid_amount"),
        (5, "NaN", "invalid_amount"),
        (5, "1e17", "invalid_amount"),
        (2, None, "missing_label"),
        (8, "=1+1", "formula_not_evaluated"),
        (8, "#DIV/0!", "spreadsheet_error"),
    ],
)
def test_bad_rows_are_disposed_not_silently_dropped(
    tmp_path: Path, column: int, value: object, reason: str
) -> None:
    row = list(ROW)
    row[column] = value
    source = tmp_path / "book.xlsx"
    digest = _source(source, [row])
    receipt = _run(source, tmp_path / "out", digest)
    assert receipt["status"] == "partial"
    assert receipt["counts"] == {
        "input": 1,
        "normalized": 0,
        "out_of_scope": 0,
        "blank": 0,
        "rejected": 1,
    }
    dispositions = pq.read_table(tmp_path / "out/row_dispositions.parquet").to_pylist()
    assert dispositions[0]["reason"] == reason
    assert dispositions[0]["source_row"] == 2
    assert dispositions[0]["source_object_sha256"] == digest
    assert dispositions[0]["record_id"] is None


def test_named_headers_and_nonpositive_amounts(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    rows = [[*ROW[:5], amount, *ROW[6:]][::-1] for amount in (0, -1, "1.125")]
    digest = _source(source, rows, list(reversed(HEADERS)))
    _run(source, tmp_path / "out", digest)
    facts = pq.read_table(tmp_path / "out/budget_facts.parquet").to_pylist()
    assert [row["amount"] for row in facts] == [
        Decimal(0),
        Decimal(-1),
        Decimal("1.125"),
    ]


@pytest.mark.parametrize(
    "headers",
    [
        HEADERS[:-2],
        [*HEADERS[:-1], "Vote"],
        [*HEADERS[:-1], None],
        [*HEADERS[:-1], 123],
    ],
)
def test_unknown_headers_fail_before_output(
    tmp_path: Path, headers: list[object]
) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source, [ROW], headers)
    with pytest.raises(ValueError, match="invalid_headers"):
        _run(source, tmp_path / "out", digest)
    assert not (tmp_path / "out").exists()


def test_hash_and_existing_output_fail_closed(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source, [ROW])
    with pytest.raises(ValueError, match="source_hash_mismatch"):
        _run(source, tmp_path / "out", "0" * 64)
    assert not (tmp_path / "out").exists()
    (tmp_path / "out").mkdir()
    with pytest.raises(FileExistsError):
        _run(source, tmp_path / "out", digest)
    assert list((tmp_path / "out").iterdir()) == []


@pytest.mark.parametrize("digest", ["a" * 63, "a" * 65, "A" * 64, "g" * 64, ""])
def test_invalid_digest_rejected_without_reading(tmp_path: Path, digest: str) -> None:
    with pytest.raises(ValueError, match="invalid_source_sha256"):
        _run(tmp_path / "absent", tmp_path / "out", digest)


@pytest.mark.parametrize(
    ("observed_at", "vintage", "locator"),
    [
        ("2026-08-30", "Budget-2025", "source"),
        ("2026-08-30T00:00:00Z", " ", "source"),
        ("2026-08-30T00:00:00Z", "Budget-2025", " "),
    ],
)
def test_invalid_context_rejected(
    tmp_path: Path, observed_at: str, vintage: str, locator: str
) -> None:
    with pytest.raises(ValueError, match="invalid_source_context"):
        normalize_budget_workbook(
            tmp_path / "absent",
            tmp_path / "out",
            expected_sha256="a" * 64,
            observed_at=observed_at,
            source_vintage=vintage,
            source_locator=locator,
        )


def test_capped_snapshot_and_boundary(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source, [ROW])
    size = source.stat().st_size
    monkeypatch.setattr(budget, "_MAX_SOURCE_BYTES", size - 1)
    with pytest.raises(ValueError, match="source_byte_limit"):
        _run(source, tmp_path / "out", digest)
    assert not (tmp_path / "out").exists()
    monkeypatch.setattr(budget, "_MAX_SOURCE_BYTES", size)
    assert _run(source, tmp_path / "out", digest)["status"] == "passed"


def test_missing_sheet_and_empty_source(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.save(source)
    workbook.close()
    with pytest.raises(ValueError, match="missing_raw_data_sheet"):
        _run(source, tmp_path / "out", hashlib.sha256(source.read_bytes()).hexdigest())
    assert not (tmp_path / "out").exists()
    digest = _source(source, [])
    receipt = _run(source, tmp_path / "out", digest)
    assert receipt["status"] == "empty"
    assert receipt["counts"] == {
        "input": 0,
        "normalized": 0,
        "out_of_scope": 0,
        "blank": 0,
        "rejected": 0,
    }
    assert pq.read_table(tmp_path / "out/budget_facts.parquet").num_rows == 0


def test_write_failure_has_no_completion_marker(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source, [ROW])

    def fail_write(*_args: object, **_kwargs: object) -> None:
        message = "injected disk failure"
        raise OSError(message)

    monkeypatch.setattr(budget.pq, "write_table", fail_write)
    with pytest.raises(OSError, match="injected disk failure"):
        _run(source, tmp_path / "out", digest)
    assert (tmp_path / "out").is_dir()
    assert not (tmp_path / "out/MANIFEST.json").exists()
    assert hashlib.sha256(source.read_bytes()).hexdigest() == digest


@pytest.mark.parametrize(
    ("column", "value", "status"),
    [
        (0, " ", "partial"),
        (0, 1, "partial"),
        (0, "#N/A", "partial"),
        (0, "health", "empty"),
        (1, None, "partial"),
        (1, 0, "partial"),
        (1, 10000, "partial"),
        (1, 1, "passed"),
        (1, 9999, "passed"),
        (5, None, "partial"),
        (5, "-1e17", "partial"),
        (5, "-Infinity", "partial"),
        (5, "99999999999999999.999", "passed"),
        (5, "-99999999999999999.999", "passed"),
        (2, " ", "partial"),
        (2, 1, "partial"),
    ],
)
def test_value_boundaries(
    tmp_path: Path, column: int, value: object, status: str
) -> None:
    row = list(ROW)
    row[column] = value
    source = tmp_path / "book.xlsx"
    digest = _source(source, [row])
    assert _run(source, tmp_path / "out", digest)["status"] == status


@pytest.mark.parametrize(
    ("rows", "status", "returncode"),
    [
        ([ROW], "passed", 0),
        ([], "empty", 2),
        ([[*ROW[:5], "bad", *ROW[6:]]], "partial", 2),
    ],
)
def test_command_reports_status(
    tmp_path: Path, rows: list[list[object]], status: str, returncode: int
) -> None:
    source = tmp_path / "book.xlsx"
    digest = _source(source, rows)
    result = subprocess.run(
        [
            sys.executable,
            "tools/build_health_budget.py",
            "--source",
            str(source),
            "--expected-sha256",
            digest,
            "--output-dir",
            str(tmp_path / "out"),
            "--observed-at",
            "2026-08-30T00:00:00Z",
            "--source-vintage",
            "Budget-2025",
            "--source-locator",
            "data/raw/b25-expenditure-data.xlsx",
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    assert result.returncode == returncode, result.stderr
    assert json.loads(result.stdout)["status"] == status
