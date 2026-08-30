"""Historical source extraction preserves annotations, periods and decimals."""

import hashlib
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook, load_workbook

from archive_govt_nz.domains.health_appropriations.historical import (
    _year_rows,
    normalize_historical_workbook,
)


def _source(
    tmp_path: Path, change: tuple[str, str, str | float | None] | None = None
) -> Path:
    book = Workbook()
    health = book.active
    assert health is not None
    health.title = "Spending"
    gdp = book.create_sheet("Nominal GDP")
    for sheet in (health, gdp):
        sheet["A3"] = "$ millions"
        for row, year in enumerate(range(1989, 1994), 5):
            sheet.cell(row, 2, year)
    health["H4"] = "Health"
    gdp["C3"] = "Nominal GDP"
    for row in range(5, 10):
        health.cell(row, 8, 605.7 if row == 5 else row)
        gdp.cell(row, 3, row * 100)
    health["A5"] = "Cash, March Years"
    health["A6"] = "Cash, June Years"
    health["A7"] = "old-GAAP"
    health["A8"] = "IFRS, June Years"
    health["A9"] = "PBE Standards, June Years"
    gdp["A5"] = " March Years"
    gdp["A6"] = "June Years"
    health["B6"] = "1990†"
    health["B7"] = "1991*"
    health["B8"] = "1992^"
    health["B9"] = "1993^#"
    for row, text in enumerate(
        ("† GST", "* GAAP caveat", "^ Restated", "# New standards"), 11
    ):
        health.cell(row, 1, text)
    health["A16"] = "% GDP"
    health["H17"] = "Health"
    health["B18"] = 1989
    health["H18"] = 0.06
    book.create_sheet("Other")["A1"] = "retained"
    if change:
        sheet, cell, value = change
        book[sheet][cell] = value
    raw = BytesIO()
    book.save(raw)
    book.close()
    result = BytesIO()
    with ZipFile(BytesIO(raw.getvalue())) as source, ZipFile(result, "w") as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b">605.7<", b">605.70000000000005<")
            target.writestr(item, payload)
    path = tmp_path / "source.xlsx"
    path.write_bytes(result.getvalue())
    return path


def _run(source: Path, out: Path) -> dict[str, object]:
    return normalize_historical_workbook(
        source,
        out,
        expected_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
        source_locator="data/raw/history.xlsx",
        source_vintage="fiscal-2024",
        observed_at="2026-08-30T00:00:00+00:00",
    )


def test_exact_values_annotations_periods_and_rebuild(tmp_path: Path) -> None:
    source = _source(tmp_path)
    before = source.read_bytes()
    receipt = _run(source, tmp_path / "one")
    assert receipt["status"] == "passed"
    facts = pq.read_table(tmp_path / "one/historical_facts.parquet").to_pylist()
    assert len(facts) == 10
    health = [row for row in facts if row["measure"] == "health_spending"]
    assert health[0]["amount"] == Decimal("605.70000000000005000")
    assert health[0]["source_number_token"] == "605.70000000000005"  # noqa: S105 - numeric source lexeme, not a credential
    assert [row["year"] for row in health] == list(range(1989, 1994))
    assert [row["period_end_month"] for row in health] == [3, 6, 6, 6, 6]
    assert health[2]["accounting_basis"] == "old-GAAP"
    assert health[4]["year_label"] == "1993^#"
    assert len(health[4]["footnotes"]) == 2
    assert all(row["valid_time_start"] is None for row in facts)
    assert str(health[0]["valid_time_end"]) == "1989-03-31"
    lineage = pq.read_table(tmp_path / "one/field_lineage.parquet").to_pylist()
    period = [
        row
        for row in lineage
        if row["record_id"] == health[2]["record_id"]
        and row["field"] == "period_end_month"
    ]
    assert period[0]["source_coordinate"] == "'Spending'!A6"
    dispositions = pq.read_table(tmp_path / "one/cell_dispositions.parquet").to_pylist()
    assert (
        next(
            row for row in dispositions if row["source_coordinate"] == "'Spending'!H18"
        )["disposition"]
        == "preserved_only"
    )
    _run(source, tmp_path / "two")
    assert source.read_bytes() == before
    for file in (tmp_path / "one").iterdir():
        assert file.read_bytes() == (tmp_path / "two" / file.name).read_bytes()
    with pytest.raises(FileExistsError):
        _run(source, tmp_path / "one")


@pytest.mark.parametrize(
    ("value", "reason"),
    [
        (None, "missing_amount"),
        ("=1+1", "formula_not_evaluated"),
        ("#VALUE!", "spreadsheet_error"),
        ("text", "non_numeric_amount"),
        (True, "non_numeric_amount"),
        (1e22, "unsupported_precision"),
    ],
)
def test_invalid_amounts_are_explicit(
    tmp_path: Path, value: str | float | None, reason: str
) -> None:
    receipt = _run(_source(tmp_path, ("Spending", "H6", value)), tmp_path / "out")
    assert receipt["status"] == "partial"
    rows = pq.read_table(tmp_path / "out/cell_dispositions.parquet").to_pylist()
    row = next(row for row in rows if row["source_coordinate"] == "'Spending'!H6")
    assert row["disposition"] == "rejected"
    assert row["reason"] == reason


@pytest.mark.parametrize(
    ("sheet", "cell", "value"),
    [
        ("Spending", "A3", "%"),
        ("Spending", "H4", "Other"),
        ("Nominal GDP", "C3", "Real GDP"),
        ("Spending", "B6", "1990?"),
        ("Spending", "B6", 1989),
        ("Spending", "B6", None),
        ("Spending", "A5", None),
        ("Spending", "A6", "Unknown"),
        ("Spending", "A11", None),
    ],
)
def test_unknown_layout_fails_without_outputs(
    tmp_path: Path, sheet: str, cell: str, value: str | float | None
) -> None:
    with pytest.raises(ValueError, match=r"historical|footnote"):
        _run(_source(tmp_path, (sheet, cell, value)), tmp_path / "out")
    assert not (tmp_path / "out").exists()


@pytest.mark.parametrize(
    ("cell", "value", "reason"),
    [
        ("A5", "old-GAAP", "historical_gaap_period"),
        ("A12", "† Duplicate", "ambiguous_footnote"),
    ],
)
def test_ambiguous_context_rejected(
    tmp_path: Path, cell: str, value: str, reason: str
) -> None:
    with pytest.raises(ValueError, match=reason):
        _run(_source(tmp_path, ("Spending", cell, value)), tmp_path / "out")


def test_empty_series_rejected() -> None:
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    with pytest.raises(ValueError, match="empty_historical_series"):
        _year_rows(sheet, 6)
    book.close()


def test_missing_sheet_rejected(tmp_path: Path) -> None:
    source = _source(tmp_path)
    book = load_workbook(source)
    book.remove(book["Spending"])
    book.save(source)
    book.close()
    with pytest.raises(ValueError, match="missing_historical_sheet"):
        _run(source, tmp_path / "out")


@pytest.mark.parametrize("value", [0, -123.5])
def test_nonpositive_source_values_preserved(tmp_path: Path, value: float) -> None:
    _run(_source(tmp_path, ("Spending", "H6", value)), tmp_path / "out")
    rows = pq.read_table(tmp_path / "out/historical_facts.parquet").to_pylist()
    fact = next(
        row
        for row in rows
        if row["measure"] == "health_spending" and row["year"] == 1990
    )
    assert fact["amount"] == Decimal(str(value))
