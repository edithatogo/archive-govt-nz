"""Health-appropriations source census and format contracts."""

from __future__ import annotations

import sqlite3
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from archive_govt_nz.domains.health_appropriations import formats
from archive_govt_nz.domains.health_appropriations.formats import (
    inventory_pdf,
    inventory_sqlite,
    inventory_workbook,
)
from archive_govt_nz.domains.health_appropriations.inventory import (
    Disposition,
    SourceInventoryRecord,
    normalize_url,
    validate_inventory,
)


def _record(source_id: str, **values: object) -> SourceInventoryRecord:
    defaults: dict[str, object] = {
        "source_id": source_id,
        "family": "treasury_vote_health",
        "title": "Vote Health",
        "url": "HTTPS://Treasury.govt.nz/item#fragment",
        "observed_at": "2026-08-29T00:00:00Z",
        "cutoff": "2026-08-29",
        "disposition": Disposition.DISCOVERED,
        "reason": "metadata observed; capture pending",
    }
    defaults.update(values)
    return SourceInventoryRecord(**defaults)  # type: ignore[arg-type]


def test_inventory_normalizes_urls_and_has_stable_identity() -> None:
    record = _record("vote-2026")
    assert record.url == "https://treasury.govt.nz/item"
    assert record.record_id == _record("vote-2026").record_id
    assert normalize_url("https://EXAMPLE.nz:443/a?q=1#x") == "https://example.nz/a?q=1"
    assert normalize_url("https://example.nz:8443/a") == "https://example.nz:8443/a"


@pytest.mark.parametrize("url", ["file:///tmp/a", "javascript:x", "https:///x"])
def test_inventory_rejects_non_public_urls(url: str) -> None:
    with pytest.raises(ValueError, match="invalid_source_url"):
        _record("bad", url=url)


def test_inventory_requires_one_disposition_and_known_predecessor() -> None:
    validate_inventory([])
    validate_inventory(
        [
            _record("old"),
            _record(
                "new",
                disposition=Disposition.SUPERSEDED,
                predecessor_source_id="old",
                reason="replacement observed",
            ),
        ]
    )
    with pytest.raises(ValueError, match="duplicate_source_disposition"):
        validate_inventory([_record("same"), _record("same")])
    with pytest.raises(ValueError, match="unknown_predecessor"):
        validate_inventory(
            [
                _record(
                    "new",
                    disposition=Disposition.SUPERSEDED,
                    predecessor_source_id="missing",
                    reason="replacement observed",
                )
            ]
        )


def test_captured_inventory_requires_sha256() -> None:
    with pytest.raises(ValueError, match="captured_source_missing_digest"):
        _record("captured", disposition=Disposition.CAPTURED, reason=None)


@pytest.mark.parametrize(
    ("values", "error"),
    [
        ({"title": ""}, "missing_source_identity"),
        ({"object_sha256": "x"}, "invalid_source_digest"),
        (
            {"disposition": Disposition.SUPERSEDED},
            "superseded_source_missing_predecessor",
        ),
        ({"reason": None}, "non_object_disposition_missing_reason"),
    ],
)
def test_inventory_rejects_incomplete_records(
    values: dict[str, object], error: str
) -> None:
    with pytest.raises(ValueError, match=error):
        _record("invalid", **values)


def test_workbook_inventory_preserves_structure(tmp_path: Path) -> None:
    path = tmp_path / "fixture.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Raw Data"
    sheet["A1"] = "Year"
    sheet["B1"] = "Amount"
    sheet["A2"] = 2026
    sheet["B2"] = "=1+2"
    sheet.merge_cells("C1:D1")
    workbook.create_sheet("Hidden").sheet_state = "hidden"
    workbook.save(path)
    before = path.read_bytes()
    result = inventory_workbook(path)
    assert result["kind"] == "xlsx"
    sheets = result["sheets"]
    assert isinstance(sheets, list)
    assert sheets[0]["formula_cells"] == 1
    assert sheets[1]["state"] == "hidden"
    assert path.read_bytes() == before


def test_workbook_inventory_rejects_traversal_member(tmp_path: Path) -> None:
    path = tmp_path / "unsafe.xlsx"
    with ZipFile(path, "w") as package:
        package.writestr("../escape", b"x")
    with pytest.raises(ValueError, match="unsafe_workbook_member"):
        inventory_workbook(path)


def test_format_inventory_rejects_invalid_packages(tmp_path: Path) -> None:
    invalid_workbook = tmp_path / "invalid.xlsx"
    invalid_workbook.write_bytes(b"not-a-zip")
    with pytest.raises(ValueError, match="invalid_workbook_package"):
        inventory_workbook(invalid_workbook)
    invalid_pdf = tmp_path / "invalid.pdf"
    invalid_pdf.write_bytes(b"not-a-pdf")
    with pytest.raises(ValueError, match="invalid_pdf"):
        inventory_pdf(invalid_pdf)


def test_workbook_inventory_enforces_package_limits(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "bounded.xlsx"
    with ZipFile(path, "w") as package:
        package.writestr("member", b"x")
    monkeypatch.setattr(formats, "_MAX_MEMBERS", 0)
    with pytest.raises(ValueError, match="workbook_member_limit"):
        inventory_workbook(path)
    monkeypatch.setattr(formats, "_MAX_MEMBERS", 20_000)
    monkeypatch.setattr(formats, "_MAX_EXPANDED_BYTES", 0)
    with pytest.raises(ValueError, match="workbook_expansion_limit"):
        inventory_workbook(path)


def test_pdf_and_sqlite_inventory(tmp_path: Path) -> None:
    pdf = tmp_path / "a.pdf"
    pdf.write_bytes(b"%PDF-1.4\n/Type /Page\n/Type /Pages\n%%EOF")
    assert inventory_pdf(pdf)["page_count"] == 1
    database = tmp_path / "a.sqlite"
    with sqlite3.connect(database) as connection:
        connection.execute("CREATE TABLE facts (value INTEGER)")
        connection.executemany("INSERT INTO facts VALUES (?)", [(1,), (2,)])
    result = inventory_sqlite(database)
    assert result["integrity"] == "ok"
    tables = result["tables"]
    assert isinstance(tables, list)
    assert tables[0]["row_count"] == 2
