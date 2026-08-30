"""Rich structural inventory never rewrites or evaluates source workbooks."""

from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from archive_govt_nz.domains.health_appropriations import formats
from archive_govt_nz.domains.health_appropriations.formats import inventory_workbook


def _write_rich_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Facts"
    sheet.append(["Year", "Amount"])
    sheet.append([2026, "=1+2"])
    sheet["A2"].comment = Comment("not exported", "not exported")
    sheet.add_table(Table(displayName="FactsTable", ref="A1:B2"))
    sheet.merge_cells("D1:E1")
    sheet.row_dimensions[2].hidden = True
    sheet.row_dimensions[3].hidden = False
    sheet.column_dimensions.group("B", "C", hidden=True)
    sheet.column_dimensions["D"].hidden = False
    workbook.defined_names.add(
        DefinedName("GlobalFacts", attr_text="'Facts'!$A$1:$B$2")
    )
    sheet.defined_names.add(DefinedName("LocalFacts", attr_text="'Facts'!$A$2"))
    chart = BarChart()
    chart.add_data(
        Reference(worksheet=sheet, min_col=2, min_row=1, max_row=2),
        titles_from_data=True,
    )
    chart.anchor = "G1"
    sheet.add_chart(chart)
    workbook.create_sheet("Hidden").sheet_state = "veryHidden"
    workbook.save(path)
    workbook.close()
    with ZipFile(path, "a") as package:
        package.writestr("custom/uninterpreted.bin", b"retained but not decoded")


def test_rich_workbook_census_retains_coordinates_and_parts(tmp_path: Path) -> None:
    path = tmp_path / "rich.xlsx"
    _write_rich_fixture(path)
    before = path.read_bytes()
    result = inventory_workbook(path)
    assert result["schema_version"] == "archive-govt-nz.workbook-inventory/v1"
    assert result["defined_names"] == ("GlobalFacts",)
    with ZipFile(path) as package:
        assert result["package_members"] == tuple(sorted(package.namelist()))
    sheets = result["sheets"]
    assert isinstance(sheets, list)
    facts = sheets[0]
    assert facts["dimension"] == "A1:E2"
    assert facts["formula_coordinates"] == ("B2",)
    assert facts["comment_coordinates"] == ("A2",)
    assert facts["merged_range_refs"] == ("D1:E1",)
    assert facts["table_ranges"] == (("FactsTable", "A1:B2"),)
    assert facts["hidden_rows"] == (2,)
    assert facts["hidden_columns"] == (("B", 2, 3),)
    assert facts["defined_names"] == ("LocalFacts",)
    assert facts["chart_count"] == 1
    assert facts["formula_cells"] == len(facts["formula_coordinates"])
    assert facts["merged_ranges"] == len(facts["merged_range_refs"])
    assert sheets[1]["state"] == "veryHidden"
    assert sheets[1]["formula_coordinates"] == ()
    assert sheets[1]["comment_coordinates"] == ()
    assert sheets[1]["hidden_rows"] == ()
    assert sheets[1]["hidden_columns"] == ()
    assert sheets[1]["table_ranges"] == ()
    assert sheets[1]["defined_names"] == ()
    assert "not exported" not in json.dumps(result)
    assert inventory_workbook(path) == result
    assert path.read_bytes() == before


def test_package_limits_accept_exact_boundary(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "exact.xlsx"
    _write_rich_fixture(path)
    with ZipFile(path) as package:
        count = len(package.infolist())
        expanded = sum(part.file_size for part in package.infolist())
    monkeypatch.setattr(formats, "_MAX_MEMBERS", count)
    monkeypatch.setattr(formats, "_MAX_EXPANDED_BYTES", expanded)
    assert inventory_workbook(path)["expanded_bytes"] == expanded
    monkeypatch.setattr(formats, "_MAX_MEMBERS", count - 1)
    with pytest.raises(ValueError, match=r"^workbook_member_limit$"):
        inventory_workbook(path)
    monkeypatch.setattr(formats, "_MAX_MEMBERS", count)
    monkeypatch.setattr(formats, "_MAX_EXPANDED_BYTES", expanded - 1)
    with pytest.raises(ValueError, match=r"^workbook_expansion_limit$"):
        inventory_workbook(path)


def test_loader_retains_formula_and_link_metadata(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "options.xlsx"
    _write_rich_fixture(path)
    original = formats.load_workbook
    views: list[bool] = []

    def load_preserving(
        filename: Path, *, read_only: bool, data_only: bool, keep_links: bool
    ) -> Workbook:
        assert not read_only
        views.append(data_only)
        assert keep_links
        return original(
            filename, read_only=read_only, data_only=data_only, keep_links=keep_links
        )

    monkeypatch.setattr(formats, "load_workbook", load_preserving)
    assert inventory_workbook(path)["kind"] == "xlsx"
    assert views == [False, True]
