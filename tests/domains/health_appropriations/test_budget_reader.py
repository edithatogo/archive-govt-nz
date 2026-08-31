"""Synthetic Budget package verification and fail-closed boundaries."""

from __future__ import annotations

import hashlib
import json
import shutil
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import Workbook

from archive_govt_nz.domains.health_appropriations import budget_reader as reader
from archive_govt_nz.domains.health_appropriations.budget import (
    normalize_budget_workbook,
)

HEADERS = (
    "Vote",
    "Year",
    "Department",
    "Appropriation Name",
    "Functional Classification",
    "Amount $000",
    "Amount Type",
    "Portfolio Name",
)
TABLES = (
    "budget_facts.parquet",
    "field_lineage.parquet",
    "row_dispositions.parquet",
)


def _pin(root: Path) -> str:
    return hashlib.sha256((root / "MANIFEST.json").read_bytes()).hexdigest()


def _manifest(root: Path, **changes: object) -> str:
    path = root / "MANIFEST.json"
    value = json.loads(path.read_text())
    value.update(changes)
    path.write_text(json.dumps(value))
    return _pin(root)


@pytest.fixture(scope="session")
def package_template(tmp_path_factory: pytest.TempPathFactory) -> Path:
    # Produce the immutable baseline once; each corruption test receives fresh
    # copies, so reader assertions remain independent without repeated XLSX IO.
    tmp_path = tmp_path_factory.mktemp("budget-reader-template")
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Raw Data"
    headers = [*HEADERS, "Extra"]
    values = [
        "Health",
        2026,
        "Health",
        "Care",
        "Health",
        -1.125,
        "Main Estimates",
        "Health",
        None,
    ]
    sheet.append(headers[::-1])
    sheet.append(values[::-1])
    sheet.append([None] * len(headers))
    sheet.append(["Education", *values[1:]][::-1])
    sheet.append(values[::-1])
    source = tmp_path / "synthetic.xlsx"
    workbook.save(source)
    workbook.close()
    root = tmp_path / "package"
    normalize_budget_workbook(
        source,
        root,
        expected_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
        observed_at="2026-08-31T12:00:00+12:00",
        source_vintage="Budget-2026",
        source_locator="synthetic.xlsx",
    )
    return root


@pytest.fixture
def package(tmp_path: Path, package_template: Path) -> Path:
    return Path(shutil.copytree(package_template, tmp_path / "package"))
def _rewrite(root: Path, name: str, rows: list[dict[str, Any]]) -> str:
    schema = pq.read_schema(root / name)
    pq.write_table(pa.Table.from_pylist(rows, schema=schema), root / name)
    hashes = json.loads((root / "MANIFEST.json").read_text())["output_sha256"]
    hashes[name] = hashlib.sha256((root / name).read_bytes()).hexdigest()
    return _manifest(root, output_sha256=hashes)


def test_valid_reordered_columns_all_rows_and_no_writes(package: Path) -> None:
    before = {path.name: path.read_bytes() for path in package.iterdir()}
    facts, lineage, dispositions, manifest = reader.read_verified_budget(
        package, _pin(package)
    )
    assert (len(facts), len(lineage), len(dispositions)) == (2, 18, 4)
    assert str(facts[0]["amount"]) == "-1.125"
    assert facts[0]["rights_state"] == manifest["rights_state"] == "not_evaluated"
    assert {path.name: path.read_bytes() for path in package.iterdir()} == before


@pytest.mark.parametrize(
    ("key", "value"),
    [
        ("status", "partial"),
        ("schema_version", "wrong"),
        ("transformation_id", "wrong"),
        ("rights_state", "approved"),
        ("counts", {}),
        ("output_sha256", {}),
        ("source_vintage", "wrong"),
        ("source_locator", "wrong"),
        ("observed_at", "2026-01-01T00:00:00Z"),
    ],
)
def test_manifest_contract(package: Path, key: str, value: object) -> None:
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _manifest(package, **{key: value}))


@pytest.mark.parametrize("value", [True, 1.0])
def test_manifest_count_types(package: Path, value: object) -> None:
    counts = json.loads((package / "MANIFEST.json").read_text())["counts"]
    assert counts["blank"] == 1
    counts["blank"] = value
    with pytest.raises(ValueError, match="budget_package_contract"):
        reader.read_verified_budget(package, _manifest(package, counts=counts))


@pytest.mark.parametrize(
    ("name", "key", "value"),
    [
        ("budget_facts.parquet", "amount", 2),
        ("budget_facts.parquet", "year", 2025),
        ("budget_facts.parquet", "record_id", "sha256:" + "0" * 64),
        ("budget_facts.parquet", "source_observation_id", "bad"),
        ("budget_facts.parquet", "lineage_id", "bad"),
        ("budget_facts.parquet", "unit", "NZD_millions"),
        ("budget_facts.parquet", "department", "other"),
        ("field_lineage.parquet", "record_id", "orphan"),
        ("field_lineage.parquet", "field", "unknown"),
        ("field_lineage.parquet", "source_coordinate", "'Other'!A2"),
        ("field_lineage.parquet", "source_coordinate", "'Raw Data'!A3"),
        ("field_lineage.parquet", "source_coordinate", "'Raw Data'!Z2"),
        ("field_lineage.parquet", "raw_value", "wrong"),
        ("field_lineage.parquet", "normalized_value", "wrong"),
        ("field_lineage.parquet", "rule", "wrong"),
        ("row_dispositions.parquet", "source_row", 3),
        ("row_dispositions.parquet", "reason", "wrong"),
        ("row_dispositions.parquet", "record_id", "wrong"),
        ("row_dispositions.parquet", "sheet", "wrong"),
        ("row_dispositions.parquet", "raw_values_json", "[]"),
        (
            "row_dispositions.parquet",
            "raw_values_json",
            '{"Vote": "Health", "Vote": "Health"}',
        ),
    ],
)
def test_rehashed_semantic_corruption(
    package: Path, name: str, key: str, value: object
) -> None:
    rows = pq.read_table(package / name).to_pylist()
    rows[0][key] = value
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _rewrite(package, name, rows))


@pytest.mark.parametrize("name", TABLES)
@pytest.mark.parametrize("operation", ["duplicate", "missing"])
def test_row_closure(package: Path, name: str, operation: str) -> None:
    rows = pq.read_table(package / name).to_pylist()
    rows = [*rows, rows[0]] if operation == "duplicate" else rows[1:]
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _rewrite(package, name, rows))


def test_same_count_duplicate_fact(package: Path) -> None:
    name = "budget_facts.parquet"
    rows = pq.read_table(package / name).to_pylist()
    assert len(rows) == 2
    rows[1] = dict(rows[0])
    with pytest.raises(ValueError, match="budget_package_contract"):
        reader.read_verified_budget(package, _rewrite(package, name, rows))


def test_cross_record_column_mapping_drift(package: Path) -> None:
    name = "field_lineage.parquet"
    rows = pq.read_table(package / name).to_pylist()
    second_record = rows[-1]["record_id"]
    entries = [row for row in rows if row["record_id"] == second_record]
    assert len(entries) == 9
    # Each record still has all fields and a valid column bijection; only
    # consistency between the two source rows has been corrupted.
    entries[0]["source_coordinate"], entries[1]["source_coordinate"] = (
        entries[1]["source_coordinate"],
        entries[0]["source_coordinate"],
    )
    with pytest.raises(ValueError, match="budget_package_contract"):
        reader.read_verified_budget(package, _rewrite(package, name, rows))


@pytest.mark.parametrize("state", ["blank", "out_of_scope"])
def test_contradictory_disposition_raw_values(package: Path, state: str) -> None:
    name = "row_dispositions.parquet"
    rows = pq.read_table(package / name).to_pylist()
    row = next(row for row in rows if row["disposition"] == state)
    raw = json.loads(row["raw_values_json"])
    raw["Extra" if state == "blank" else "Vote"] = "Health"
    row["raw_values_json"] = json.dumps(raw)
    with pytest.raises(ValueError, match="budget_package_contract"):
        reader.read_verified_budget(package, _rewrite(package, name, rows))


@pytest.mark.parametrize("vote", [None, "Education", "Health"])
def test_writer_empty_status_and_explicit_zero(
    tmp_path: Path, vote: str | None
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Raw Data"
    sheet.append(list(HEADERS))
    if vote is not None:
        sheet.append([vote, 2026, "Health", "Care", "Health", 0, "Actuals", "Health"])
    source = tmp_path / "source.xlsx"
    workbook.save(source)
    workbook.close()
    root = tmp_path / "package"
    receipt = normalize_budget_workbook(
        source,
        root,
        expected_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
        observed_at="2026-08-31T00:00:00Z",
        source_vintage="synthetic",
        source_locator="synthetic",
    )
    if vote == "Health":
        assert receipt["status"] == "passed"
        facts, lineage, dispositions, _ = reader.read_verified_budget(root, _pin(root))
        assert len(facts) == len(dispositions) == 1
        assert len(lineage) == 8
        assert facts[0]["amount"] == 0
    else:
        # Writer-produced empty packages are legitimate outputs, but outside
        # the reader's reviewed passed/nonempty package consumption contract.
        assert receipt["status"] == "empty"
        with pytest.raises(ValueError, match="budget_package_contract"):
            reader.read_verified_budget(root, _pin(root))


@pytest.mark.parametrize(
    "limit", ["MAX_BYTES", "MAX_TOTAL_BYTES", "MAX_ROWS", "MAX_EXPANDED_BYTES"]
)
def test_exact_resource_boundaries(
    package: Path, monkeypatch: pytest.MonkeyPatch, limit: str
) -> None:
    sizes = [path.stat().st_size for path in package.iterdir()]
    row_counts, expanded = [], []
    for name in TABLES:
        metadata = pq.read_metadata(package / name)
        row_counts.append(metadata.num_rows)
        expanded.append(
            sum(
                metadata.row_group(i).total_byte_size
                for i in range(metadata.num_row_groups)
            )
        )
    boundary = {
        "MAX_BYTES": max(sizes),
        "MAX_TOTAL_BYTES": sum(sizes),
        "MAX_ROWS": max(row_counts),
        "MAX_EXPANDED_BYTES": max(expanded),
    }[limit]
    monkeypatch.setattr(reader, limit, boundary)
    reader.read_verified_budget(package, _pin(package))
    monkeypatch.setattr(reader, limit, boundary - 1)
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _pin(package))


def test_pin_file_set_symlinks_and_duplicate_json(
    package: Path, tmp_path: Path
) -> None:
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, "invalid")
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, "0" * 64)
    extra = package / "extra"
    extra.touch()
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _pin(package))
    extra.unlink()
    link = tmp_path / "link"
    link.symlink_to(package, target_is_directory=True)
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(link, _pin(package))
    original = package / "MANIFEST.json"
    copied = tmp_path / "manifest-copy"
    original.rename(copied)
    original.symlink_to(copied)
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _pin(package))


def test_schema_mismatch(package: Path) -> None:
    name = "budget_facts.parquet"
    pq.write_table(pa.table({"wrong": [1]}), package / name)
    hashes = json.loads((package / "MANIFEST.json").read_text())["output_sha256"]
    hashes[name] = hashlib.sha256((package / name).read_bytes()).hexdigest()
    with pytest.raises(
        ValueError,
        match=r"budget_package_contract|source_hash_mismatch|source_byte_limit",
    ):
        reader.read_verified_budget(package, _manifest(package, output_sha256=hashes))


def test_final_pin_recheck(package: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    original = reader.verified_snapshot
    calls = []

    def change_manifest(path: Path, digest: str, *, max_bytes: int) -> bytes:
        if path.name == "MANIFEST.json":
            if calls:
                _manifest(package, source_vintage="changed")
            calls.append(path)
        return original(path, digest, max_bytes=max_bytes)

    monkeypatch.setattr(reader, "verified_snapshot", change_manifest)
    with pytest.raises(ValueError, match="source_hash_mismatch"):
        reader.read_verified_budget(package, _pin(package))


@pytest.mark.parametrize(
    ("key", "value"),
    [("max_row", 1), ("max_row", 100002), ("max_column", 7), ("max_column", 16385)],
)
def test_inventory_bounds(package: Path, key: str, value: int) -> None:
    inventory = json.loads((package / "MANIFEST.json").read_text())[
        "workbook_inventory"
    ]
    inventory["sheets"][0][key] = value
    with pytest.raises(ValueError, match="budget_package_contract"):
        reader.read_verified_budget(
            package, _manifest(package, workbook_inventory=inventory)
        )


@settings(max_examples=8, deadline=None)
@given(
    year=st.integers(min_value=1, max_value=9999),
    amount=st.integers(min_value=-1000000, max_value=1000000),
)
def test_generated_exact_values(year: int, amount: int) -> None:
    # Workbook/Parquet filesystem latency is not a semantic property deadline.
    with TemporaryDirectory() as directory:
        base = Path(directory)
        source = base / "book.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Raw Data"
        sheet.append(list(HEADERS))
        sheet.append(
            [
                "Health",
                year,
                "Department",
                "Care",
                "Health",
                amount,
                "Actuals",
                "Health",
            ]
        )
        workbook.save(source)
        workbook.close()
        root = base / "out"
        normalize_budget_workbook(
            source,
            root,
            expected_sha256=hashlib.sha256(source.read_bytes()).hexdigest(),
            observed_at="2026-08-31T00:00:00Z",
            source_vintage="synthetic",
            source_locator="synthetic",
        )
        facts, lineage, dispositions, _ = reader.read_verified_budget(root, _pin(root))
        assert len(facts) == len(dispositions) == 1
        assert len(lineage) == 8
        assert facts[0]["year"] == year
        assert facts[0]["amount"] == amount
